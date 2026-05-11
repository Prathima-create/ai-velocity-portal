#!/usr/bin/env python3
"""
FinCom QC Automation Dashboard - v4
Outputs (per process+month folder):
  - QC_Dashboard_<folder>.html   (interactive dashboard, share via SharePoint)
  - QC_Report_<folder>.xlsx       (formal Excel report for leadership)
  - QC_Snapshot_<folder>.png      (screenshot to paste in email body)

Key rules:
  - Process file = source for ALL org-level numbers
  - Analyst file = source for analyst-level rows ONLY
  - Defect cell rule: string == "0" => defect (letter "O" is NOT a defect)
  - Defects KPI = sum of "# of Missed Parameters" column
  - Fatal: "Appropriate Resolution" or "Confidentiality" == "0"
  - Non-Fatal Defect: any param == "0" but neither fatal param == "0"
"""

import os, sys, json, base64, warnings
warnings.filterwarnings('ignore')
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter, defaultdict

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas not installed. Run: python -m pip install pandas --user")
    sys.exit(1)


# ============================================================
# CONSTANTS
# ============================================================
DESKTOP     = Path.home() / "Desktop"
BASE_FOLDER = DESKTOP / "FinCom_QC"

PARAMETERS = [
    "1. Greeting", "2. Empathy", "3. Language",
    "4. Appropriate Resolution", "5. Delay/Escalation",
    "6. Phone Calls", "7. Proactive/Self-Service",
    "8. Transfer/SIM CTI", "9. Confidentiality",
    "10. Annotations", "11. Resolution Code",
]
FATAL_NAMES = {"appropriateresolution", "confidentiality"}

# Defect Impact Categorization — maps each parameter (normalized name) to a business-impact category
# Categories: Vendor Experience, Process Compliance, Controllership, Financial Impact, Accounting
IMPACT_CATEGORY_MAP = {
    "greeting":              "Vendor Experience",
    "empathy":               "Vendor Experience",
    "language":              "Vendor Experience",
    "appropriateresolution": "Vendor Experience",
    "delayescalation":       "Vendor Experience",
    "phonecalls":            "Vendor Experience",
    "proactiveselfservice":  "Vendor Experience",
    "transfersimcti":        "Process Compliance",
    "confidentiality":       "Controllership",
    "annotations":           "Process Compliance",
    "resolutioncode":        "Process Compliance",
}
IMPACT_CATEGORIES_ORDER = ["Vendor Experience", "Process Compliance", "Controllership", "Financial Impact", "Accounting"]
IMPACT_CATEGORY_DEFS = {
    "Vendor Experience":  "Impacting the vendor/customer",
    "Process Compliance": "Process / SOP not followed",
    "Controllership":     "Impacting Amazon's internal and external policies",
    "Financial Impact":   "Substantial loss to amazon or vendor by short or excess payment",
    "Accounting":         "Accounting errors impacting P&L and Balance sheet",
}
MONTHS_ORDER = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


# ============================================================
# HELPERS
# ============================================================
def norm(s):
    return (str(s).strip().lower()
            .replace('/', '').replace('-', '')
            .replace('_', '').replace(' ', '')
            .replace('.', '').replace(',', ''))


def is_defect(cell):
    return str(cell).strip() == "0"


def is_fatal_param(param_name):
    n = norm(param_name)
    return any(fk in n for fk in FATAL_NAMES)


def find_col_strict(df, *candidates):
    """Strict column lookup. Exact match, then prefix match (with non-alphanumeric boundary)."""
    cols = list(df.columns)
    for cand in candidates:
        key = norm(cand)
        if not key:
            continue
        for col in cols:
            if col == key:
                return col
        for col in cols:
            if col.startswith(key):
                rest = col[len(key):]
                if not rest or not rest[0].isalnum():
                    return col
        for col in cols:
            if len(col) >= 3 and key.startswith(col):
                return col
    return None


def get_series(df, *candidates):
    col = find_col_strict(df, *candidates)
    if col is not None:
        return df[col].fillna('').astype(str).str.strip()
    return pd.Series([''] * len(df), index=df.index)


def parse_accuracy_series(series):
    def _p(v):
        try:
            v = float(str(v).replace('%','').strip())
            return round(v * 100, 2) if v <= 1.0 else round(v, 2)
        except Exception:
            return None
    return series.apply(_p)


def parse_date_series(series):
    def _p(v):
        v = str(v).strip().split(' ')[0].split(',')[0]
        if not v or v.lower() in ('nan','none','','na','n/a'):
            return None
        parts = v.replace('-', '/').split('/')
        if len(parts) != 3:
            return None
        try:
            a, b, c = int(parts[0]), int(parts[1]), int(parts[2])
            year = c if c > 1000 else 2000 + c
            if a > 12: return datetime(year, b, a)
            if b > 12: return datetime(year, a, b)
            return datetime(year, a, b)
        except Exception:
            return None
    return series.apply(_p)


def iso_week(d):
    if d is None or pd.isna(d):
        return None
    try:
        return f"W{d.isocalendar()[1]}"
    except Exception:
        return None


def workdays_between(start, end):
    if start is None or end is None:
        return None
    if end < start:
        return 0
    days = 0
    cur = start
    while cur <= end:
        if cur.weekday() < 5:
            days += 1
        cur += timedelta(days=1)
    return max(0, days - 1)


# ============================================================
# CSV LOADER + CONFIG
# ============================================================
def load_csv_smart(filepath, required_keywords=None):
    if not filepath.exists():
        return None, f"File not found: {filepath.name}"
    try:
        raw = pd.read_csv(filepath, header=None, encoding='utf-8-sig',
                          on_bad_lines='skip', dtype=str)
    except Exception as e:
        return None, str(e)

    header_idx = 0
    if required_keywords:
        found = False
        for i in range(min(20, len(raw))):
            row = raw.iloc[i]
            normalized = [norm(v) for v in row.values if pd.notna(v) and str(v).strip()]
            if len(normalized) < 3:  # title rows have only 1-2 cells
                continue
            if all(any(kw in cell for cell in normalized) for kw in required_keywords):
                header_idx = i
                found = True
                break
        if not found:
            for i in range(min(20, len(raw))):
                row = raw.iloc[i]
                non_empty = [v for v in row.values if pd.notna(v) and str(v).strip()]
                if len(non_empty) >= 3:
                    header_idx = i
                    break

    df = pd.read_csv(filepath, header=header_idx, encoding='utf-8-sig',
                     on_bad_lines='skip', dtype=str)
    df.columns = [norm(c) for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated(keep='first')]
    df = df.dropna(how='all')
    return df, None


def load_config(folder):
    config_path = folder / "config.csv"
    defaults = {
        "disputes_sla_days":      7,
        "rectification_sla_days": 2,
        "ivoc_sla_days":          2,
        "reduction_sla_days":     5,
        "accuracy_target":        95.0,
        "aged":     0,
        "resolved": 0,
        "reopen":   0,
        "hmd":      0,
        "total":    0,
        "defect_reduction_target_pct": 10.0,
        "fatal_inap_position": 4,   # 1-based parameter position for Inappropriate Resolution
        "fatal_conf_position": 9,   # 1-based parameter position for Confidentiality
    }
    if not config_path.exists():
        print("  WARNING: config.csv not found — using defaults")
        return defaults

    df, err = load_csv_smart(config_path)
    if err or df is None or len(df) == 0:
        return defaults

    row = df.iloc[0]
    def _get(*keys):
        for k in keys:
            col = find_col_strict(df, k)
            if col:
                try:
                    val = str(row[col]).strip().replace(',', '')
                    # Strip "days" / "%" / etc to extract number
                    import re
                    m = re.search(r'-?\d+\.?\d*', val)
                    if m:
                        return float(m.group())
                except Exception:
                    pass
        return None

    cfg = dict(defaults)
    # SLA values
    for k in ['disputes_sla_days', 'rectification_sla_days', 'ivoc_sla_days', 'reduction_sla_days']:
        v = _get(k, k.replace('_', ' '), k.replace('_sla_days', ' SLA'))
        if v is not None and v > 0:
            cfg[k] = v
    # Inflow categories
    for k in ['aged', 'resolved', 'reopen', 'hmd', 'total']:
        v = _get(k)
        if v is not None and v >= 0:
            cfg[k] = v
    # Defect reduction target
    v = _get('defect_reduction_target_pct', 'defect reduction target', 'dr target', 'reduction target')
    if v is not None and v > 0:
        cfg['defect_reduction_target_pct'] = v
    # Fatal positions
    v = _get('fatal_inap_position', 'fatal inap position', 'inap position')
    if v is not None and v >= 1:
        cfg['fatal_inap_position'] = int(v)
    v = _get('fatal_conf_position', 'fatal conf position', 'conf position')
    if v is not None and v >= 1:
        cfg['fatal_conf_position'] = int(v)

    # Parameters list (comma-separated). Default = General's 11. SHT can override with 10 in different order.
    cfg['parameters'] = list(PARAMETERS)  # default
    params_col = find_col_strict(df, 'Parameters', 'Parameter List', 'Params')
    if params_col:
        try:
            params_str = str(row[params_col]).strip()
            if params_str and params_str.lower() != 'nan':
                # Split on comma, strip each
                parsed = [p.strip() for p in params_str.split(',') if p.strip()]
                if len(parsed) >= 1:
                    # Number them so they match PARAMETERS format ("1. Greeting" etc.)
                    cfg['parameters'] = [f"{i+1}. {name}" for i, name in enumerate(parsed)]
                    print(f"  Config: Custom parameter list with {len(parsed)} params loaded")
        except Exception as e:
            print(f"  WARNING: Could not parse Parameters column: {e}")
    print(f"  Config: Fatal Inap @ position {cfg['fatal_inap_position']}, Fatal Conf @ position {cfg['fatal_conf_position']}")

    # Compute total if missing
    if cfg['total'] == 0:
        cfg['total'] = cfg['aged'] + cfg['resolved'] + cfg['reopen'] + cfg['hmd']

    return cfg


# ============================================================
# AUDIT FILE PARSER (Process or Analyst)
#
# Real Fincom audit files have a complex structure:
#   - 60+ rows of preamble (definitions, instructions)
#   - True header row contains: Analyst, ORG, Audit Date, Accuracy %, # of Missed Parameters
#   - Parameter columns come AFTER "# of Missed Parameters" in pairs:
#     [point_value, Comment, point_value, Comment, ...]
#   - Header for parameter columns is the POINT VALUE (3, 8, 15, "O", 22, etc.)
#   - Cell value: point value = passed; "0" = defect; "O" = not applicable
#
# Strategy: find the real header row, then read 11 parameters by POSITION
# (alternating param-col, comment-col) starting after "# of Missed Parameters".
# ============================================================
def parse_audit_file(filepath, source_label, parameters=None, inap_pos=4, conf_pos=9):
    """Parse audit CSV. parameters: optional custom list (defaults to global PARAMETERS).
    inap_pos / conf_pos: 1-based parameter positions for fatal columns."""
    if parameters is None:
        parameters = list(PARAMETERS)
    num_params = len(parameters)
    if not filepath.exists():
        print(f"  WARNING: {filepath.name} not found")
        return pd.DataFrame()

    print(f"\n  --- {filepath.name} ---")

    # Read raw file to find the header row
    try:
        raw = pd.read_csv(filepath, header=None, encoding='utf-8-sig',
                          on_bad_lines='skip', dtype=str)
    except Exception as e:
        print(f"  ERROR: Could not load: {e}")
        return pd.DataFrame()

    # Find the header row.
    # STRICT REQUIREMENTS (avoid false-match in preamble/instructions):
    # - Each required keyword must appear as ITS OWN cell (not part of a longer sentence)
    # - The matching row must have many non-empty cells (real headers have ~40+ columns,
    #   preamble rows are usually sparse or have long instruction sentences in few cells)
    header_row_idx = None
    for i in range(min(200, len(raw))):
        row = raw.iloc[i]
        cells = [str(v).strip() for v in row.values if pd.notna(v) and str(v).strip()]
        if len(cells) < 10:  # real header has many columns
            continue
        # Each cell must be reasonably short (header cells aren't long sentences)
        # Allow up to 60 chars for cells like "Case Date (Resolved/ Reopen/ HMD Poll)"
        normalized = [norm(c) for c in cells if len(c) <= 60]
        # Each keyword must MATCH a cell (not just be a substring of a long sentence)
        has_analyst  = any(c == 'analyst' or c == 'primaryanalyst' for c in normalized)
        has_accuracy = any(c == 'accuracy' or c == 'accuracypct' or 'accuracy' in c and len(c) <= 15 for c in normalized)
        has_date     = any(c == 'auditdate' or c == 'date' for c in normalized)
        has_case     = any(c == 'casenumber' or c == 'caseno' or c == 'caseid' for c in normalized)
        if has_analyst and has_accuracy and has_date and has_case:
            header_row_idx = i
            break

    if header_row_idx is None:
        print("  ERROR: Could not find header row containing Analyst + Accuracy + Audit Date + Case Number")
        return pd.DataFrame()

    print(f"  Header row found at line {header_row_idx + 1}")

    # Re-read with the correct header row
    df = pd.read_csv(filepath, header=header_row_idx, encoding='utf-8-sig',
                     on_bad_lines='skip', dtype=str)
    df = df.dropna(how='all')

    # Get raw column names BEFORE normalization (needed for positional logic)
    raw_cols = list(df.columns)
    df.columns = [norm(c) for c in df.columns]
    # NOTE: Do NOT dedup columns here - it shifts positions and breaks positional reads.
    # Duplicate columns are tolerated; we use df.iloc[:, pos] for reads.

    # Find the column index of "# of Missed Parameters" — params start right after
    missed_idx = None
    for i, c in enumerate(df.columns):
        if 'missedparameters' in c or 'ofmissed' in c or c == 'missed':
            missed_idx = i
            break

    if missed_idx is None:
        print("  WARNING: '# of Missed Parameters' column not found — parameter detection may fail")
        missed_idx = len(df.columns)  # fallback

    # The 11 parameters live at columns: missed_idx+1, +3, +5, +7, +9, +11, +13, +15, +17, +19, +21
    # (every other column after Missed; the columns in between are 'Comment' columns)
    param_col_positions = []
    for i in range(num_params):
        pos = missed_idx + 1 + (i * 2)
        if pos < len(df.columns):
            param_col_positions.append(pos)

    print(f"  Found Missed Params col at index {missed_idx}; reading {num_params} parameters at positions {param_col_positions}")

    # CRITICAL FIX: identify fatal columns by DESCRIPTION TEXT in raw row 1 (parameter title row),
    # not by position. SHT and General may have different column orders.
    # NOTE: text-match override DISABLED - was causing regressions in General.
    # If SHT has different column ordering, that's a separate config we'll address later.
    fatal_overrides = {}

    # Use config-driven positions for fatal parameters
    # inap_pos and conf_pos are 1-based parameter positions (e.g., 4 means the 4th parameter column)
    print(f"  Using fatal positions from config: Inap @ #{inap_pos}, Conf @ #{conf_pos}")

    # Map each parameter NAME to a column POSITION (integer index)
    # We use positions, not names, because SHT files have duplicate header values
    # (e.g., multiple columns named "5" or "o1") which break name-based pandas lookup
    param_positions = {}
    for i, p_name in enumerate(parameters):
        if i < len(param_col_positions):
            param_positions[p_name] = param_col_positions[i]
        else:
            param_positions[p_name] = None

    # OVERRIDE fatal parameter positions based on text-match if found
    for p_name in parameters:
        np_key = norm(p_name)
        for fkey, ci in fatal_overrides.items():
            if fkey in np_key:
                old_pos = param_positions.get(p_name)
                if old_pos != ci:
                    print(f"  OVERRIDE: '{p_name}' position-based was col {old_pos} -> text-matched to col {ci}")
                param_positions[p_name] = ci

    found = sum(1 for c in param_positions.values() if c is not None)
    print(f"  Parameter columns matched: {found}/{len(parameters)}")
    for p, ci in param_positions.items():
        marker = "OK " if ci is not None else "MISS"
        col_name = df.columns[ci] if ci is not None and ci < len(df.columns) else 'N/A'
        print(f"     {marker}  {p:32s} -> col index {ci} ('{col_name}')")

    # Filter rows: keep only those with a real Case Number (numeric)
    case_col = find_col_strict(df, "Case Number", "Case No", "Case ID")
    if case_col:
        df = df[df[case_col].astype(str).str.strip().str.match(r'^\d+(-\d+)?$', na=False)].copy()
        print(f"  After filtering to numeric Case Numbers: {len(df)} rows")

    if len(df) == 0:
        print("  ERROR: No data rows found after filtering")
        return pd.DataFrame()

    # Basic fields
    df['_source']  = source_label
    df['_analyst'] = get_series(df, "Analyst")
    df['_org']     = get_series(df, "ORG", "Org").str.strip().str.upper().replace('', 'Unknown')
    df['_case']    = get_series(df, "Case Number", "Case No", "Case ID")
    df['_topic']   = get_series(df, "Issue", "Topic")
    df['_category']= get_series(df, "Case Category", "Category")
    df['_manager'] = get_series(df, "Manager Login", "Manager")
    # Comment column for fatal drill-downs — use the right-most "Comment" column near the end,
    # or fall back to any "Comment" column. Real file has many "Comment" columns (one per param).
    # Best match for a single audit comment: a column named "Audit Comment" or the last
    # standalone "Comment" column. We'll just use the get_series helper which finds the first.
    df['_comment'] = get_series(df, "Audit Comment", "Comment", "Auditor Remark", "Notes")

    missed = pd.to_numeric(get_series(df, "# of Missed Parameters", "Missed Parameters", "# of Missed", "Missed"),
                           errors='coerce').fillna(0).astype(int)
    df['_missed'] = missed

    df['_accuracy'] = parse_accuracy_series(get_series(df, "Accuracy %", "Accuracy"))

    df['_date'] = parse_date_series(get_series(df, "Audit Date", "Date"))
    df['_week'] = df['_date'].apply(iso_week)

    # Per-row defect detection (string equality)
    # Build a map of parameter -> ADJACENT comment column POSITION
    param_comment_positions = {}
    for p_name, p_pos in param_positions.items():
        if p_pos is not None and p_pos + 1 < len(df.columns):
            param_comment_positions[p_name] = p_pos + 1
        else:
            param_comment_positions[p_name] = None

    # Build name-keyed param_cols dict (using positions to extract column names)
    # but use df.iloc[:, position] inside the row loop to avoid duplicate-column ambiguity
    fatal_inap, fatal_conf, nonfatal, hits_list, fatal_comments = [], [], [], [], []
    for row_idx in range(len(df)):
        hits = []
        is_inap = is_conf = False
        comments_for_row = []
        for p in parameters:
            pos = param_positions[p]
            if pos is None:
                continue
            # Use df.iloc[row, pos] for safe positional access regardless of duplicate column names
            try:
                cell = df.iloc[row_idx, pos]
            except (IndexError, KeyError):
                continue
            if is_defect(cell):
                hits.append(p)
                if "appropriateresolution" in norm(p):
                    is_inap = True
                if "confidentiality" in norm(p):
                    is_conf = True
                # Pick up the comment in the adjacent column for this defect
                cpos = param_comment_positions.get(p)
                if cpos is not None:
                    try:
                        cval = str(df.iloc[row_idx, cpos]).strip()
                        if cval and cval.lower() != 'nan':
                            comments_for_row.append(cval)
                    except (IndexError, KeyError):
                        pass
        hits_list.append(hits)
        fatal_inap.append(is_inap)
        fatal_conf.append(is_conf)
        nonfatal.append(len(hits) > 0 and not (is_inap or is_conf))
        # Concatenate all defect comments for this row (joined with " | ")
        fatal_comments.append(" | ".join(comments_for_row) if comments_for_row else "")

    df['_param_hits']  = hits_list
    df['_fatal_inap']  = fatal_inap
    df['_fatal_conf']  = fatal_conf
    df['_is_fatal']    = pd.Series(fatal_inap, index=df.index) | pd.Series(fatal_conf, index=df.index)
    df['_is_nonfatal'] = nonfatal
    df['_has_any_def'] = df['_param_hits'].apply(lambda x: len(x) > 0)
    # Use the per-defect-column comment as the "audit comment" for drill-downs
    # If empty, fall back to the generic comment field (rare)
    df['_comment'] = pd.Series(fatal_comments, index=df.index).where(
        pd.Series(fatal_comments, index=df.index) != '', df['_comment']
    )

    print(f"  {filepath.name}: {len(df)} records loaded")
    keep = ['_source','_analyst','_org','_case','_topic','_category','_manager',
            '_comment','_missed','_accuracy','_date','_week',
            '_param_hits','_fatal_inap','_fatal_conf',
            '_is_fatal','_is_nonfatal','_has_any_def']
    return df[keep].copy()


# ============================================================
# DISPUTES
# ============================================================
def parse_disputes(filepath, sla_days=7):
    if not filepath.exists():
        return pd.DataFrame()
    df, err = load_csv_smart(filepath, ["case"])
    if err or df is None:
        return pd.DataFrame()

    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    case_col = find_col_strict(df, "Case Number", "Case No", "Case ID")
    if case_col:
        df = df[df[case_col].astype(str).str.strip().str.match(r'^\d+(-\d+)?$', na=False)].copy()

    if len(df) == 0:
        print(f"  Disputes: 0 records (empty file)")
        return pd.DataFrame()

    df['_case']    = get_series(df, "Case Number", "Case No", "Case ID")
    df['_owner']   = get_series(df, "Defect Owner", "Owner").str.lower()
    df['_report']  = get_series(df, "Report Information", "Dispute Outcome", "Outcome").str.lower()
    df['_org']     = get_series(df, "ORG", "Org").str.strip().str.upper().replace('', 'Unknown')
    df['_analyst'] = get_series(df, "Analyst", "Primary Analyst")
    df['_manager'] = get_series(df, "Manager Login", "Manager")

    df['_audit_date']   = parse_date_series(get_series(df, "Audit Date"))
    df['_dispute_date'] = parse_date_series(get_series(df, "Day of Dispute", "Date of Dispute", "Dispute Date", "Date"))

    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    def _scalar(v):
        if hasattr(v, 'iloc'):
            return v.iloc[0] if len(v) > 0 else None
        return v

    df['_workdays'] = df.apply(
        lambda r: workdays_between(_scalar(r['_audit_date']), _scalar(r['_dispute_date'])), axis=1)

    def _sla(row):
        ad = _scalar(row['_audit_date']); dd = _scalar(row['_dispute_date'])
        try:
            if dd is None or ad is None or pd.isna(dd) or pd.isna(ad):
                return "Pending"
        except Exception:
            return "Pending"
        wd = _scalar(row['_workdays'])
        if wd is None: return "Pending"
        return "SLA Met" if wd <= sla_days else "SLA Breached"
    sla_result = df.apply(_sla, axis=1)
    if isinstance(sla_result, pd.DataFrame):
        sla_result = sla_result.iloc[:, 0]
    df['_sla'] = sla_result

    def _category(row):
        owner = str(_scalar(row['_owner'])).lower()
        report = str(_scalar(row['_report'])).lower()
        if 'auditor' in owner: return "QC Error (Auditor)"
        if 'quest lead' in owner or 'questlead' in owner: return "To Quest Lead"
        if 'backup' in owner or 'back up' in owner: return "Moved to Backup"
        if 'primary' in owner: return "Stayed with Primary"
        if 'reverse' in report: return "Reversed"
        if 'reject' in report: return "Rejected"
        return "Other"
    cat_result = df.apply(_category, axis=1)
    if isinstance(cat_result, pd.DataFrame):
        cat_result = cat_result.iloc[:, 0]
    df['_category'] = cat_result

    print(f"  Disputes: {len(df)} records loaded")
    return df


# ============================================================
# RECTIFICATION + IVOC (two-state)
# ============================================================
def _two_state_parser(filepath, label, sla_days, status_keys, end_keys):
    """3-state parser: Rectified / Pending / Not Actioned (ops miss)"""
    if not filepath.exists():
        return pd.DataFrame()
    df, err = load_csv_smart(filepath, ["case"])
    if err or df is None:
        return pd.DataFrame()

    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    case_col = find_col_strict(df, "Case Number", "Case No", "Case ID")
    if case_col:
        df = df[df[case_col].astype(str).str.strip().str.match(r'^\d+(-\d+)?$', na=False)].copy()

    # Empty data file - return empty df gracefully (status card will show zeros)
    if len(df) == 0:
        print(f"  {label}: 0 records (empty file)")
        return pd.DataFrame()

    df['_case']    = get_series(df, "Case Number", "Case No", "Case ID")
    df['_org']     = get_series(df, "ORG", "Org").str.strip().str.upper().replace('', 'Unknown')
    df['_analyst'] = get_series(df, "Analyst", "Primary Analyst")
    df['_manager'] = get_series(df, "Manager Login", "Manager")
    df['_status_raw'] = get_series(df, *status_keys)
    df['_action']     = get_series(df, "Action Taken", "Action", "Feedback Provided")

    df['_audit_date'] = parse_date_series(get_series(df, "Audit Date", "Paste Date"))
    df['_end_date']   = parse_date_series(get_series(df, *end_keys))

    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    def _scalar(v):
        if hasattr(v, 'iloc'):
            return v.iloc[0] if len(v) > 0 else None
        return v

    df['_workdays'] = df.apply(
        lambda r: workdays_between(_scalar(r['_audit_date']), _scalar(r['_end_date'])), axis=1)

    def _status(row):
        end_date = _scalar(row['_end_date'])
        audit_date = _scalar(row['_audit_date'])
        action = str(_scalar(row['_action'])).lower().strip()
        status_raw = str(_scalar(row['_status_raw'])).lower().strip()

        # Rectified: end date exists OR explicit "rectif" status OR action documented
        try:
            if end_date is not None and pd.notna(end_date):
                return "Rectified"
        except Exception:
            pass
        if 'rectif' in status_raw and 'not' not in status_raw:
            return "Rectified"
        if action and action not in ('na','n/a','none','pending','no action',''):
            return "Rectified"
        # No action yet - check if past SLA window
        try:
            if audit_date is not None and pd.notna(audit_date):
                wd_today = workdays_between(audit_date, datetime.now())
                if wd_today is not None and wd_today > sla_days:
                    return "Pending — Missed SLA"
        except Exception:
            pass
        return "Pending — Within SLA"

    status_result = df.apply(_status, axis=1)
    if isinstance(status_result, pd.DataFrame):
        status_result = status_result.iloc[:, 0]
    df['_status'] = status_result

    def _sla(row):
        end_date = _scalar(row['_end_date'])
        audit_date = _scalar(row['_audit_date'])
        wd = _scalar(row['_workdays'])
        try:
            if end_date is None or pd.isna(end_date):
                if audit_date is not None and pd.notna(audit_date):
                    wd_today = workdays_between(audit_date, datetime.now())
                    if wd_today is not None and wd_today > sla_days:
                        return "SLA Breached"
                return "Pending"
        except Exception:
            return "Pending"
        if wd is None: return "Pending"
        return "SLA Met" if wd <= sla_days else "SLA Breached"

    sla_result = df.apply(_sla, axis=1)
    if isinstance(sla_result, pd.DataFrame):
        sla_result = sla_result.iloc[:, 0]
    df['_sla'] = sla_result

    print(f"  {label}: {len(df)} records loaded")
    return df


def parse_rectification(filepath, sla_days=5):
    return _two_state_parser(filepath, "Rectification", sla_days,
        status_keys=["Rectification Status", "Status"],
        end_keys=["Rectification Date", "Closed Date", "Date"])


def parse_ivoc(filepath, sla_days=5):
    return _two_state_parser(filepath, "IVOC", sla_days,
        status_keys=["IVOC Status", "Rectification Status", "Status"],
        end_keys=["IVOC Rectification Date", "Rectification Date", "Closed Date", "Date"])


# ============================================================
# DEFECT REDUCTION (filtered to current month)
# ============================================================
def parse_defect_reduction(filepath, current_month_short, sla_days=5):
    """Defect Reduction tracker.
    Reads Status column (Completed / Not Started) — drives state.
    SLA: Case Update Date -> 5 working days for ops to action.
    3 states: Action Taken / Pending — Within SLA / Pending — Missed SLA
    Skips phantom (blank) rows automatically by filtering on valid Case Number."""
    if not filepath.exists():
        return pd.DataFrame()
    df, err = load_csv_smart(filepath, ["analyst"])
    if err or df is None:
        return pd.DataFrame()

    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    # Filter out phantom rows: must have a valid Case Number (numeric)
    case_col = find_col_strict(df, "Case Number", "Case No", "Case ID")
    if case_col:
        df = df[df[case_col].astype(str).str.strip().str.match(r'^\d+(-\d+)?$', na=False)].copy()

    if len(df) == 0:
        print(f"  Defect Reduction: 0 records (empty file)")
        return pd.DataFrame()

    df['_case']    = get_series(df, "Case Number", "Case No", "Case ID")
    df['_org']     = get_series(df, "ORG", "Org").str.strip().str.upper().replace('', 'Unknown')
    df['_analyst'] = get_series(df, "Analyst")
    df['_manager'] = get_series(df, "Manager Login", "Manager")
    df['_action']  = get_series(df, "Action Taken", "Action", "Feedback Provided")
    df['_topic']   = get_series(df, "Topic", "Defect Parameter", "Parameter")
    df['_status_raw'] = get_series(df, "Status")  # NEW: read the Status column

    df['_case_update_date'] = parse_date_series(get_series(df, "Case Update Date", "Update Date"))
    df['_review_date']      = parse_date_series(get_series(df, "Date of Review", "Review Date"))

    df = df.loc[:, ~df.columns.duplicated(keep='first')]

    def _scalar(v):
        if hasattr(v, 'iloc'):
            return v.iloc[0] if len(v) > 0 else None
        return v

    def _status(row):
        status = str(_scalar(row['_status_raw'])).lower().strip()
        # Completed = action taken
        if 'complet' in status or 'done' in status or 'closed' in status:
            return "Action Taken"
        # Not Started = check SLA window
        cud = _scalar(row['_case_update_date'])
        try:
            if cud is not None and pd.notna(cud):
                wd_today = workdays_between(cud, datetime.now())
                if wd_today is not None and wd_today > sla_days:
                    return "Pending — Missed SLA"
        except Exception:
            pass
        return "Pending — Within SLA"

    status_result = df.apply(_status, axis=1)
    if isinstance(status_result, pd.DataFrame):
        status_result = status_result.iloc[:, 0]
    df['_status'] = status_result

    def _sla(row):
        s = _scalar(row['_status'])
        if s == "Action Taken":
            return "SLA Met"
        if s == "Pending — Missed SLA":
            return "SLA Breached"
        return "Pending"

    sla_result = df.apply(_sla, axis=1)
    if isinstance(sla_result, pd.DataFrame):
        sla_result = sla_result.iloc[:, 0]
    df['_sla'] = sla_result

    print(f"  Defect Reduction: {len(df)} records loaded")
    return df

# ============================================================
# METRICS COMPUTATION
# Process file -> all org-level numbers
# Analyst file -> analyst rows only
# ============================================================
def compute_metrics(process_df, analyst_df,
                    disputes_df, rect_df, ivoc_df, defred_df,
                    prev_process_df, prev_analyst_df, config, prev_config,
                    process_label, month_label, prev_month_label):
    """Build all dashboard numbers."""
    if process_df is None or len(process_df) == 0:
        return None

    m = {}

    # ---- HEADLINE KPIs (Process file = primary) ----
    total_audited = len(process_df)
    accuracy_avg  = round(process_df['_accuracy'].mean(), 2) if total_audited else 0.0

    # Process file fatal counts (primary)
    inap_count = int(process_df['_fatal_inap'].sum())
    conf_count = int(process_df['_fatal_conf'].sum())
    fatal_count = inap_count + conf_count

    # Analyst file fatal counts (parallel measure)
    if analyst_df is not None and len(analyst_df) > 0 and '_fatal_inap' in analyst_df.columns:
        analyst_inap = int(analyst_df['_fatal_inap'].sum())
        analyst_conf = int(analyst_df['_fatal_conf'].sum())
        analyst_fatal = analyst_inap + analyst_conf
    else:
        analyst_inap = analyst_conf = analyst_fatal = None

    # Defects = SUM of "# of Missed Parameters" column
    defects_total = int(process_df['_missed'].sum())
    nonfatal_defects = int(process_df['_is_nonfatal'].sum())

    m['headline'] = {
        'audited':         total_audited,
        'accuracy':        accuracy_avg,
        'fatal':           fatal_count,
        'inap':            inap_count,
        'conf':            conf_count,
        'analyst_fatal':   analyst_fatal,
        'analyst_inap':    analyst_inap,
        'analyst_conf':    analyst_conf,
        'defects':         defects_total,
        'nonfatal_defects': nonfatal_defects,
        'defect_rate':     round(defects_total / total_audited * 100, 2) if total_audited else 0,
    }

    # ---- STATUS CARDS ----
    def _status_card(df, status_field, expected_states=None):
        """Returns card data even for empty df (shows zeros for transparency).
        expected_states: list of state names to ensure all are shown even if 0."""
        states = expected_states or []
        if df is None or len(df) == 0:
            breakdown = {s: 0 for s in states}
            return {
                'total': 0,
                'breakdown': breakdown,
                'sla_met': 0,
                'sla_total': 0,
                'sla_pct': None,
            }
        breakdown = df[status_field].value_counts().to_dict() if status_field in df.columns else {}
        # Ensure all expected states show up (with 0 if missing)
        for s in states:
            if s not in breakdown:
                breakdown[s] = 0
        sla_counts = df['_sla'].value_counts().to_dict() if '_sla' in df.columns else {}
        sla_met = sla_counts.get('SLA Met', 0)
        sla_total = len(df)
        return {
            'total':     sla_total,
            'breakdown': breakdown,
            'sla_met':   sla_met,
            'sla_total': sla_total,
            'sla_pct':   round(sla_met / sla_total * 100, 2) if sla_total else None,
        }

    DISPUTE_CATS = ['Stayed with Primary', 'Moved to Backup', 'QC Error (Auditor)', 'To Quest Lead']
    THREE_STATES = ['Rectified', 'Pending — Within SLA', 'Pending — Missed SLA']
    DR_STATES    = ['Action Taken', 'Pending — Within SLA', 'Pending — Missed SLA']

    m['disputes']         = _status_card(disputes_df, '_category', DISPUTE_CATS)
    m['rectification']    = _status_card(rect_df, '_status', THREE_STATES)
    m['ivoc']             = _status_card(ivoc_df, '_status', THREE_STATES)
    m['defect_reduction'] = _status_card(defred_df, '_status', DR_STATES)

    # Compute Reversed vs Not Reversed for Disputes
    # Reversed = Defect Owner contains "auditor" (QC error). Everything else = Not Reversed.
    if disputes_df is not None and len(disputes_df) > 0 and '_owner' in disputes_df.columns:
        owner_series = disputes_df['_owner'].astype(str).str.lower()
        reversed_count = int(owner_series.str.contains('auditor', na=False).sum())
        not_reversed_count = len(disputes_df) - reversed_count
        total_disp = len(disputes_df)
        m['disputes']['reversed'] = reversed_count
        m['disputes']['not_reversed'] = not_reversed_count
        m['disputes']['reversed_pct'] = round(reversed_count / total_disp * 100, 1) if total_disp else 0
        m['disputes']['not_reversed_pct'] = round(not_reversed_count / total_disp * 100, 1) if total_disp else 0
    else:
        m['disputes']['reversed'] = 0
        m['disputes']['not_reversed'] = 0
        m['disputes']['reversed_pct'] = None
        m['disputes']['not_reversed_pct'] = None

    # ---- WoW (Cases Audited / Fatal Defects / Non-Fatal Defects per ISO week) ----
    wow_data = []
    if total_audited:
        wk_groups = process_df.groupby('_week')
        weeks_sorted = sorted(
            [w for w in wk_groups.groups.keys() if w and w != 'Unknown'],
            key=lambda w: int(w.replace('W', '')) if w.startswith('W') else 0
        )
        prev_fatal = None
        prev_nonfatal = None
        for w in weeks_sorted:
            g = wk_groups.get_group(w)
            audited = len(g)
            # Fatal = sum of fatal marks (Inap + Conf) for this week
            fatal_w = int(g['_fatal_inap'].sum()) + int(g['_fatal_conf'].sum())
            nonfatal_w = int(g['_is_nonfatal'].sum())
            wow_data.append({
                'week':      w,
                'audited':   audited,
                'fatal':     fatal_w,
                'nonfatal':  nonfatal_w,
                'fatal_delta_arrow':    None if prev_fatal is None else ('up' if fatal_w > prev_fatal else ('down' if fatal_w < prev_fatal else 'flat')),
                'nonfatal_delta_arrow': None if prev_nonfatal is None else ('up' if nonfatal_w > prev_nonfatal else ('down' if nonfatal_w < prev_nonfatal else 'flat')),
            })
            prev_fatal = fatal_w
            prev_nonfatal = nonfatal_w
    m['wow'] = wow_data

    # ---- MoM (current month vs previous month) ----
    mom = None
    if prev_process_df is not None and len(prev_process_df) > 0:
        cur_aud = total_audited
        cur_fat = fatal_count
        cur_proc_acc = accuracy_avg  # Process accuracy (current)
        cur_anl_acc = round(analyst_df['_accuracy'].mean(), 2) if analyst_df is not None and len(analyst_df) > 0 else None

        prv_aud = len(prev_process_df)
        prv_fat = int(prev_process_df['_fatal_inap'].sum()) + int(prev_process_df['_fatal_conf'].sum())
        prv_proc_acc = round(prev_process_df['_accuracy'].mean(), 2) if prv_aud else 0.0
        prv_anl_acc = round(prev_analyst_df['_accuracy'].mean(), 2) if prev_analyst_df is not None and len(prev_analyst_df) > 0 else None

        mom = {
            'prev_label': prev_month_label or 'Prev',
            'cur_label':  month_label,
            'prev_audited': prv_aud, 'cur_audited': cur_aud,
            'prev_fatal':   prv_fat, 'cur_fatal':   cur_fat,
            'prev_proc_acc': prv_proc_acc, 'cur_proc_acc': cur_proc_acc,
            'prev_anl_acc':  prv_anl_acc,  'cur_anl_acc':  cur_anl_acc,
            'prev_fatal_pct': round(prv_fat / prv_aud * 100, 2) if prv_aud else 0,
            'cur_fatal_pct':  round(cur_fat / cur_aud * 100, 2) if cur_aud else 0,
            'fatal_delta':    cur_fat - prv_fat,
            'audited_delta':  cur_aud - prv_aud,
            'proc_acc_delta': round(cur_proc_acc - prv_proc_acc, 2),
            'anl_acc_delta':  round((cur_anl_acc or 0) - (prv_anl_acc or 0), 2) if (cur_anl_acc is not None and prv_anl_acc is not None) else None,
        }
    m['mom'] = mom

    # ---- INFLOW MoM (compare config Resolved/HMD/Reopen Apr vs Mar) ----
    inflow_rows = []
    if prev_config:
        for cat_key, cat_label in [('resolved','Resolved'), ('hmd','HMD'), ('reopen','Reopen')]:
            cur_v = config.get(cat_key, 0)
            prv_v = prev_config.get(cat_key, 0)
            if prv_v > 0:
                pct = round((cur_v - prv_v) / prv_v * 100, 1)
            else:
                pct = None
            flag = abs(pct) >= 50 if pct is not None else False
            inflow_rows.append({
                'category': cat_label,
                'prev': int(prv_v),
                'cur': int(cur_v),
                'delta_pct': pct,
                'flag': flag,
            })
    m['inflow_mom'] = inflow_rows

    # ---- AUDIT COVERAGE (audited / config inflow) ----
    coverage_rows = []
    # Count audited cases per category from process file
    cat_col = '_category' if '_category' in process_df.columns else None
    cat_counts = {}
    if cat_col:
        for v in process_df[cat_col].fillna('').astype(str).str.strip().str.lower():
            for k in ['aged', 'resolved', 'reopen', 'hmd']:
                if k in v:
                    cat_counts[k] = cat_counts.get(k, 0) + 1
                    break
    for k, label in [('aged','Aged'), ('resolved','Resolved'), ('reopen','Reopen'), ('hmd','HMD')]:
        total = config.get(k, 0)
        audited = cat_counts.get(k, 0)
        cov = round(audited / total * 100, 1) if total > 0 else None
        coverage_rows.append({
            'category': label,
            'total': int(total),
            'audited': int(audited),
            'coverage_pct': cov,
        })
    m['coverage'] = coverage_rows

    # ---- DEFECT REDUCTION GOAL ----
    # Inappropriate Resolution count: April vs March (from process files)
    target_pct = config.get('defect_reduction_target_pct', 10.0)
    cur_inap = int(process_df['_fatal_inap'].sum())
    prv_inap = int(prev_process_df['_fatal_inap'].sum()) if prev_process_df is not None and len(prev_process_df) > 0 else None
    if prv_inap is not None and prv_inap > 0:
        actual_change_pct = round((cur_inap - prv_inap) / prv_inap * 100, 1)
        # Goal met if reduced by AT LEAST target%
        # i.e. cur <= prv * (1 - target/100)
        goal_met = actual_change_pct <= -target_pct
    else:
        actual_change_pct = None
        goal_met = None
    m['defect_reduction_goal'] = {
        'target_pct': target_pct,
        'prev_count': prv_inap,
        'cur_count': cur_inap,
        'actual_change_pct': actual_change_pct,
        'goal_met': goal_met,
        'prev_label': prev_month_label or 'Prev',
        'cur_label': month_label,
    }

    # ---- TOP DEFECT PARAMETERS (all 11, ranked) ----
    counter = Counter()
    for hits in process_df['_param_hits']:
        for p in hits:
            counter[p] += 1
    m['top_defects'] = counter.most_common(len(PARAMETERS))

    # ---- IMPACT CATEGORIZATION ----
    # Group defect counts by business-impact category (Vendor Experience / Process Compliance / etc.)
    impact_counts = {cat: 0 for cat in IMPACT_CATEGORIES_ORDER}
    for param_name, cnt in m['top_defects']:
        np = norm(param_name)
        # Find matching category in map by checking if any key is a substring
        cat = None
        for k, v in IMPACT_CATEGORY_MAP.items():
            if k in np:
                cat = v
                break
        if cat is None:
            cat = "Process Compliance"  # safe default
        impact_counts[cat] += cnt
    m['impact_categorization'] = [
        {'category': cat, 'definition': IMPACT_CATEGORY_DEFS[cat], 'defects': impact_counts[cat]}
        for cat in IMPACT_CATEGORIES_ORDER
    ]

    # ---- ORG TABLE (sortable) ----
    org_rows = []
    for org, g in process_df.groupby('_org'):
        if not org:
            continue
        oa = len(g)
        # Fatal = simple sum of fatal marks (Inap + Conf), matching headline KPI
        of = int(g['_fatal_inap'].sum()) + int(g['_fatal_conf'].sum())
        # Non-fatal defects = cases with at least one non-fatal defect (no fatal mark)
        on = int(g['_is_nonfatal'].sum())
        org_rows.append({
            'org':         org,
            'audited':     oa,
            'accuracy':    round(g['_accuracy'].mean(), 2) if oa else 0,
            'fatal':       of,
            'nonfatal':    on,
            'defect_rate': round((of + on) / oa * 100, 2) if oa else 0,
        })
    m['orgs'] = sorted(org_rows, key=lambda x: -x['audited'])

    # ---- FATAL BY ANALYST (from ANALYST file only) ----
    if analyst_df is not None and len(analyst_df) > 0:
        # Sum of fatal marks (Inap + Conf) per analyst — matches KPI counting
        analyst_df_copy = analyst_df.copy()
        analyst_df_copy['_fatal_marks'] = analyst_df_copy['_fatal_inap'].astype(int) + analyst_df_copy['_fatal_conf'].astype(int)
        fba = (analyst_df_copy[analyst_df_copy['_fatal_marks'] > 0]
               .groupby('_analyst')['_fatal_marks'].sum()
               .sort_values(ascending=False).head(10))
        m['fatal_by_analyst'] = [(a, int(c)) for a, c in fba.items() if a]
    else:
        m['fatal_by_analyst'] = []

    # ---- ANALYST TABLE (from ANALYST file only) ----
    analysts = []
    if analyst_df is not None and len(analyst_df) > 0:
        for analyst, g in analyst_df.groupby('_analyst'):
            if not analyst:
                continue
            aa = len(g)
            # Fatal = sum of fatal marks (Inap + Conf)
            af = int(g['_fatal_inap'].sum()) + int(g['_fatal_conf'].sum())
            an = int(g['_is_nonfatal'].sum())
            analysts.append({
                'analyst':  analyst,
                'org':      g['_org'].mode().iloc[0] if not g['_org'].mode().empty else '',
                'manager':  g['_manager'].mode().iloc[0] if not g['_manager'].mode().empty else '',
                'audited':  aa,
                'accuracy': round(g['_accuracy'].mean(), 2),
                'fatal':    af,
                'nonfatal': an,
            })
    m['analysts'] = sorted(analysts, key=lambda x: x['accuracy'])

    # ---- FILTER VALUES ----
    m['filters'] = {
        'orgs':       sorted({o for o in process_df['_org'].unique() if o}),
        'analysts':   sorted({a for a in process_df['_analyst'].unique() if a}),
        'categories': sorted({c for c in process_df['_category'].unique() if c}),
        'weeks':      sorted({w for w in process_df['_week'].unique() if w},
                             key=lambda w: int(w.replace('W','')) if w and w.startswith('W') else 0),
        'topics':     sorted({t for t in process_df['_topic'].unique() if t}),
    }

    # ---- ROW-LEVEL DATA for client-side drill-downs ----
    rows = []
    for _, r in process_df.iterrows():
        rows.append({
            'case':        r['_case'],
            'analyst':     r['_analyst'],
            'org':         r['_org'],
            'manager':     r['_manager'],
            'category':    r['_category'],
            'topic':       r['_topic'],
            'week':        r['_week'] or '',
            'date':        r['_date'].strftime('%Y-%m-%d') if r['_date'] is not None else '',
            'accuracy':    r['_accuracy'],
            'is_fatal':    bool(r['_is_fatal']),
            'is_nonfatal': bool(r['_is_nonfatal']),
            'has_def':     bool(r['_has_any_def']),
            'missed':      int(r['_missed']),
            'params':      r['_param_hits'],
            'comment':     r['_comment'],
        })
    m['rows'] = rows
    return m


# ============================================================
# HTML DASHBOARD BUILDER
# ============================================================
def rag_class(pct, *, inverted=False, sla_card=False):
    if pct is None:
        return 'rag-na'
    if sla_card:
        return 'rag-green' if pct >= 100 else 'rag-red'
    if inverted:
        if pct <= 5:  return 'rag-green'
        if pct <= 10: return 'rag-amber'
        return 'rag-red'
    if pct >= 95: return 'rag-green'
    if pct >= 90: return 'rag-amber'
    return 'rag-red'


def fmt_pct(p):
    return '—' if p is None else f'{p:.1f}%'


def build_html(metrics, process_label, month_label, prev_month_label, missing_files):
    h = metrics['headline']

    # KPI ROW
    acc_init_cls = rag_class(h['accuracy'])
    def_init_cls = ''
    kpi_html = f'''
    <div class="row kpi-row">
      <div class="kpi-card" onclick="openDrill('audited')">
        <div class="kpi-label">Cases Audited</div>
        <div class="kpi-value" id="kpi-audited">{h['audited']}</div>
        <div class="kpi-sub" id="kpi-audited-sub">From Process file</div>
      </div>
      <div class="kpi-card {acc_init_cls}" id="kpi-accuracy-card" onclick="openDrill('accuracy')">
        <div class="kpi-label">Accuracy</div>
        <div class="kpi-value" id="kpi-accuracy">{fmt_pct(h['accuracy'])}</div>
        <div class="kpi-sub">Average across all audits</div>
      </div>
      <div class="kpi-card kpi-fatal" onclick="openDrill('fatal')">
        <div class="kpi-label">Fatal Errors</div>
        <div class="kpi-value" id="kpi-fatal">{h['fatal']}{f" / {h['analyst_fatal']}" if h['analyst_fatal'] is not None else ''}</div>
        <div class="kpi-sub" id="kpi-fatal-sub">Process: Inap {h['inap']} | Conf {h['conf']}{f"  •  Analyst: Inap {h['analyst_inap']} | Conf {h['analyst_conf']}" if h['analyst_fatal'] is not None else ''}</div>
      </div>
      <div class="kpi-card" id="kpi-defects-card" onclick="openDrill('defects')">
        <div class="kpi-label">Defects</div>
        <div class="kpi-value" id="kpi-defects">{h['defects']}</div>
        <div class="kpi-sub" id="kpi-defects-sub">Total missed parameters</div>
      </div>
    </div>'''

    # STATUS ROW
    def _status_html(title, data, key, extra_rows=''):
        if data is None:
            return ''
        rows_inner = ''
        for label, count in data['breakdown'].items():
            label_safe = str(label).replace("'", "&#39;")
            rows_inner += f'<div class="status-row" onclick="event.stopPropagation();openStatusDrill(\'{key}\',\'{label_safe}\')"><span>{label_safe}</span><strong>{count}</strong></div>'
        # Handle SLA display for empty card
        if data['sla_total'] == 0:
            sla_display = 'SLA: N/A'
            sla_cls = 'rag-na'
        else:
            sla_cls = rag_class(data['sla_pct'], sla_card=True)
            sla_display = f"SLA: {data['sla_met']}/{data['sla_total']} ({fmt_pct(data['sla_pct'])})"
        return f'''
        <div class="status-card" id="sec-status-{key}" onclick="openDrill('{key}')">
          <div class="status-header">{title}<button class="dl-btn" style="float:right; margin-top:-2px" onclick="event.stopPropagation();downloadSectionCSV('sec-status-{key}','{title.replace(' ','_')}')">⬇</button></div>
          <div class="status-total">Total: <strong>{data['total']}</strong></div>
          <div class="status-body">{rows_inner}{extra_rows}</div>
          <div class="status-sla {sla_cls}">{sla_display}</div>
        </div>'''

    # Build disputes extra rows for Reversed / Not Reversed
    d = metrics['disputes']
    disputes_extra = ''
    if d.get('total', 0) > 0:
        disputes_extra = f'''
          <div class="status-divider"></div>
          <div class="status-row"><span>Reversed (QC Error)</span><strong>{d.get('reversed', 0)} ({d.get('reversed_pct', 0):.1f}%)</strong></div>
          <div class="status-row"><span>Not Reversed</span><strong>{d.get('not_reversed', 0)} ({d.get('not_reversed_pct', 0):.1f}%)</strong></div>
        '''

    files_present = metrics.get('files_present', {
        'disputes': True, 'rectification': True, 'ivoc': True, 'defect_reduction': True
    })

    # Count how many status cards will render
    status_card_count = sum([
        1 if files_present.get('disputes', True) else 0,
        1 if files_present.get('rectification', True) else 0,
        1 if files_present.get('ivoc', True) else 0,
        1 if files_present.get('defect_reduction', True) else 0,
    ])

    status_cards_html = ''
    if files_present.get('disputes', True):
        status_cards_html += _status_html('Disputes Status', metrics['disputes'], 'disputes', disputes_extra)
    if files_present.get('rectification', True):
        status_cards_html += _status_html('Rectification Status', metrics['rectification'], 'rectification')
    if files_present.get('ivoc', True):
        status_cards_html += _status_html('IVOC Status', metrics['ivoc'], 'ivoc')
    if files_present.get('defect_reduction', True):
        status_cards_html += _status_html('Defect Reduction Status', metrics['defect_reduction'], 'defect_reduction')

    # Defer the actual layout decision until after top_html is built (later in this function)
    # For now just store the raw cards HTML; we'll wrap it later
    status_html = f'<div class="row status-row-block">{status_cards_html}</div>'

    # ---- Inflow MoM + Audit Coverage + DR Goal (3 inline sections) ----
    inflow_html = ''
    if metrics.get('inflow_mom'):
        rows = ''
        for r in metrics['inflow_mom']:
            delta_str = f"{r['delta_pct']:+.1f}%" if r['delta_pct'] is not None else '—'
            flag_str = ' ⚠️' if r['flag'] else ''
            row_cls = 'inflow-flag' if r['flag'] else ''
            rows += f'<tr class="{row_cls}"><td>{r["category"]}</td><td>{r["prev"]:,}</td><td>{r["cur"]:,}</td><td>{delta_str}{flag_str}</td></tr>'
        inflow_html = f'''
        <div class="section third" id="sec-inflow"><div class="section-header"><h2>Inflow MoM</h2><button class="dl-btn" onclick="downloadSectionCSV('sec-inflow', 'Inflow_MoM')">Download</button></div>
          <table class="data-table">
            <thead><tr><th>Category</th><th>Prev</th><th>Cur</th><th>Δ</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
          <p class="muted small">Source: config.csv. ⚠️ = Δ ≥ 50%</p>
        </div>'''

    coverage_html = ''
    if metrics.get('coverage'):
        rows = ''
        for r in metrics['coverage']:
            cov_str = f"{r['coverage_pct']:.1f}%" if r['coverage_pct'] is not None else '—'
            rows += f'<tr><td>{r["category"]}</td><td>{r["total"]:,}</td><td>{r["audited"]:,}</td><td>{cov_str}</td></tr>'
        coverage_html = f'''
        <div class="section third" id="sec-coverage"><div class="section-header"><h2>Audit Coverage</h2><button class="dl-btn" onclick="downloadSectionCSV('sec-coverage', 'Audit_Coverage')">Download</button></div>
          <table class="data-table">
            <thead><tr><th>Category</th><th>Inflow</th><th>Audited</th><th>%</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </div>'''

    goal_html = ''
    g = metrics.get('defect_reduction_goal')
    if g and g.get('actual_change_pct') is not None and files_present.get('defect_reduction', True):
        if g['goal_met']:
            badge = '<span class="goal-met">✅ Met</span>'
        else:
            badge = '<span class="goal-miss">❌ Not Met</span>'
        goal_html = f'''
        <div class="section third" id="sec-goal"><div class="section-header"><h2>Defect Reduction Goal</h2><button class="dl-btn" onclick="downloadSectionCSV('sec-goal', 'Defect_Reduction_Goal')">Download</button></div>
          <p class="goal-line">Target: <strong>−{g['target_pct']:.0f}%</strong> MoM (Inappropriate Resolution)</p>
          <p class="goal-line">{g['prev_label']}: <strong>{g['prev_count']}</strong> → {g['cur_label']}: <strong>{g['cur_count']}</strong></p>
          <p class="goal-line">Actual: <strong>{g['actual_change_pct']:+.1f}%</strong></p>
          <p class="goal-line">{badge}</p>
        </div>'''

    # Analyst Pareto section
    pareto_html = ''
    if metrics.get('analysts'):
        analysts_with_def = [{'name': a['analyst'], 'org': a['org'],
                              'def_count': a['fatal'] + a['nonfatal']}
                             for a in metrics['analysts']]
        analysts_with_def = [a for a in analysts_with_def if a['def_count'] > 0]
        total_def_count = sum(a['def_count'] for a in analysts_with_def)
        if total_def_count > 0 and len(analysts_with_def) >= 5:
            top_n = 8 if len(analysts_with_def) >= 10 else 5
            top_analysts = sorted(analysts_with_def, key=lambda x: -x['def_count'])[:top_n]
            top_count = sum(a['def_count'] for a in top_analysts)
            top_pct = top_count / total_def_count * 100
            rows_inner = ''
            for a in top_analysts:
                pct = a['def_count'] / total_def_count * 100
                rows_inner += f'<tr><td>{a["name"]}</td><td>{a["org"]}</td><td>{a["def_count"]}</td><td>{pct:.1f}%</td></tr>'
            pareto_html = f'''
            <div class="section half" id="sec-pareto"><div class="section-header"><h2>Analyst Pareto</h2><button class="dl-btn" onclick="downloadSectionCSV('sec-pareto', 'Analyst_Pareto')">Download</button></div>
              <p class="goal-line"><strong>Top {top_n} analysts contributed {top_pct:.0f}% of defects ({top_count} of {total_def_count})</strong></p>
              <table class="data-table">
                <thead><tr><th>Analyst</th><th>Org</th><th>Defects</th><th>% of Total</th></tr></thead>
                <tbody>{rows_inner}</tbody>
              </table>
              <p class="muted small">Sorted by defect contribution. Use for coaching prioritization.</p>
            </div>'''

    extra_row = ''
    if inflow_html or coverage_html or goal_html:
        extra_row = f'<div class="row">{inflow_html}{coverage_html}{goal_html}</div>'

    # WoW
    wow_html = '<div class="section half" id="sec-wow"><div class="section-header"><h2>Week over Week</h2><button class="dl-btn" onclick="downloadSectionCSV(\'sec-wow\', \'Week_over_Week\')">Download</button></div>'
    if metrics['wow']:
        for metric_key, metric_label, color in [
            ('audited', 'Cases Audited', '#3b82f6'),
            ('fatal', 'Fatal Defects', '#dc2626'),
            ('nonfatal', 'Non-Fatal Defects', '#f59e0b'),
        ]:
            max_val = max(w[metric_key] for w in metrics['wow']) or 1
            wow_html += f'<div class="wow-block"><div class="wow-title">{metric_label}</div><div class="wow-bars">'
            for w in metrics['wow']:
                pct = w[metric_key] / max_val * 100
                arrow = ''
                if metric_key in ('fatal','nonfatal'):
                    a = w.get(f'{metric_key}_delta_arrow')
                    if a == 'up':   arrow = '<span class="arr-up">↑</span>'
                    elif a == 'down': arrow = '<span class="arr-dn">↓</span>'
                wow_html += f'''<div class="wow-bar-col">
                  <div class="wow-num">{w[metric_key]} {arrow}</div>
                  <div class="wow-bar"><div class="wow-fill" style="height:{pct}%;background:{color}"></div></div>
                  <div class="wow-lbl">{w['week']}</div>
                </div>'''
            wow_html += '</div></div>'
    else:
        wow_html += '<p class="muted">No weekly data.</p>'
    wow_html += '</div>'

    # MoM
    mom_html = '<div class="section half" id="sec-mom"><div class="section-header"><h2>Month over Month</h2><button class="dl-btn" onclick="downloadSectionCSV(\'sec-mom\', \'Month_over_Month\')">Download</button></div>'
    mom = metrics['mom']
    if mom:
        prev = prev_month_label or 'Previous'
        cur  = month_label
        def _delta(d, suffix='', better_is_lower=False):
            if d == 0:
                return '<span class="muted">no change</span>'
            if better_is_lower:
                cls = 'arr-dn' if d < 0 else 'arr-up'
                arrow = '↓' if d < 0 else '↑'
            else:
                cls = 'arr-up' if d > 0 else 'arr-dn'
                arrow = '↑' if d > 0 else '↓'
            return f'<span class="{cls}">{arrow} {abs(d)}{suffix} vs {prev}</span>'
        # Helper: format an accuracy MoM cell (handle None for missing files)
        def _acc_cell(prev_v, cur_v, delta_v):
            prev_str = fmt_pct(prev_v) if prev_v is not None else '—'
            cur_str  = fmt_pct(cur_v)  if cur_v  is not None else '—'
            delta_str = _delta(delta_v, '%') if delta_v is not None else ''
            return prev_str, cur_str, delta_str

        proc_prev, proc_cur, proc_delta = _acc_cell(mom['prev_proc_acc'], mom['cur_proc_acc'], mom['proc_acc_delta'])
        anl_prev,  anl_cur,  anl_delta  = _acc_cell(mom['prev_anl_acc'],  mom['cur_anl_acc'],  mom['anl_acc_delta'])

        mom_html += f'''
        <table class="mom-table">
          <tr><th></th><th>{prev}</th><th></th><th>{cur}</th></tr>
          <tr>
            <td class="mom-lbl">Total Audited</td>
            <td class="mom-val">{mom['prev_audited']}</td>
            <td class="vs">vs</td>
            <td class="mom-val">{mom['cur_audited']}<div class="mom-delta">{_delta(mom['audited_delta'])}</div></td>
          </tr>
          <tr>
            <td class="mom-lbl">Fatal</td>
            <td class="mom-val">{mom['prev_fatal']} ({mom['prev_fatal_pct']:.1f}%)</td>
            <td class="vs">vs</td>
            <td class="mom-val mom-fatal" onclick="openMoMDrill('fatal')">{mom['cur_fatal']} ({mom['cur_fatal_pct']:.1f}%)<div class="mom-delta">{_delta(mom['fatal_delta'], '', better_is_lower=True)}</div></td>
          </tr>
          <tr>
            <td class="mom-lbl">Process Accuracy</td>
            <td class="mom-val">{proc_prev}</td>
            <td class="vs">vs</td>
            <td class="mom-val" onclick="openMoMDrill('proc_accuracy')">{proc_cur}<div class="mom-delta">{proc_delta}</div></td>
          </tr>
          <tr>
            <td class="mom-lbl">Analyst Accuracy</td>
            <td class="mom-val">{anl_prev}</td>
            <td class="vs">vs</td>
            <td class="mom-val" onclick="openMoMDrill('anl_accuracy')">{anl_cur}<div class="mom-delta">{anl_delta}</div></td>
          </tr>
        </table>
        <p class="muted small">Click <strong>Fatal</strong> or <strong>Accuracy</strong> on the right to drill down.</p>
        '''
    else:
        mom_html += f'<p class="muted">Previous month folder not found. To enable MoM, ensure the previous month\'s data exists at <code>{BASE_FOLDER}\\{process_label}_&lt;PrevMonth&gt;_&lt;Year&gt;</code>.</p>'
    mom_html += '</div>'

    # IMPACT CATEGORIZATION SECTION
    impact_html = ''
    if metrics.get('impact_categorization'):
        impact_html = '<div class="section half" id="sec-impact"><div class="section-header"><h2>Defect Impact Categorization</h2><button class="dl-btn" onclick="downloadSectionCSV(\'sec-impact\', \'Defect_Impact_Categorization\')">Download</button></div>'
        impact_html += '<table class="data-table impact-table"><thead><tr><th>Category</th><th>Defects</th></tr></thead><tbody>'
        for r in metrics['impact_categorization']:
            impact_html += f'<tr><td>{r["category"]}</td><td>{r["defects"]}</td></tr>'
        impact_html += '</tbody></table></div>'

    # TOP DEFECTS (all 11, ranked)
    # If only 1 status card is shown, top_defects becomes a sibling in that row → use 'section' (full width when alone, half when paired)
    top_section_class = "section half"
    top_html = f'<div class="{top_section_class}" id="sec-topdefects"><div class="section-header"><h2>Top Defect Parameters</h2><button class="dl-btn" onclick="downloadSectionCSV(\'sec-topdefects\', \'Top_Defects\')">Download</button></div>'
    if metrics['top_defects']:
        max_c = metrics['top_defects'][0][1] if metrics['top_defects'] else 1
        top_html += '<div class="bar-chart">'
        for name, count in metrics['top_defects']:
            pct = (count / max_c * 100) if max_c else 0
            name_safe = str(name).replace("'", "&#39;")
            top_html += f'''
            <div class="bar-row" onclick="openParamDrill('{name_safe}')">
              <div class="bar-label">{name_safe}</div>
              <div class="bar-track"><div class="bar-fill" style="width:{pct}%"></div></div>
              <div class="bar-value">{count}</div>
            </div>'''
        top_html += '</div>'
    else:
        top_html += '<p class="muted">No defects recorded.</p>'
    top_html += '</div>'

    # Top Defects always renders in its own row, paired with Impact Categorization
    top_html_inline = top_html

    # ORG TABLE (sortable)
    org_html = '<div class="section half" id="sec-org"><div class="section-header"><h2>Org Scorecard</h2><button class="dl-btn" onclick="downloadSectionCSV(\'sec-org\', \'Org_Scorecard\')">Download</button></div>'
    org_html += '''<table class="data-table" id="orgTable">
      <thead><tr>
        <th onclick="sortDataTable('orgTable',0,'str')">Org</th>
        <th onclick="sortDataTable('orgTable',1,'num')">Cases Audited</th>
        <th onclick="sortDataTable('orgTable',2,'num')">Accuracy</th>
        <th onclick="sortDataTable('orgTable',3,'num')">Fatal</th>
        <th onclick="sortDataTable('orgTable',4,'num')">Non-Fatal Defects</th>
      </tr></thead><tbody>'''
    for o in metrics['orgs']:
        acc_cls = rag_class(o['accuracy'])
        org_html += f'''<tr onclick="openOrgDrill('{o['org']}')">
          <td><strong>{o['org']}</strong></td>
          <td>{o['audited']}</td>
          <td class="{acc_cls}">{fmt_pct(o['accuracy'])}</td>
          <td>{o['fatal']}</td>
          <td>{o['nonfatal']}</td>
        </tr>'''
    org_html += '</tbody></table></div>'

    # FATAL BY ANALYST
    fba_html = '<div class="section half" id="sec-fba"><div class="section-header"><h2>Fatal Errors by Analyst</h2><button class="dl-btn" onclick="downloadSectionCSV(\'sec-fba\', \'Fatal_by_Analyst\')">Download</button></div>'
    if metrics['fatal_by_analyst']:
        max_c = metrics['fatal_by_analyst'][0][1] if metrics['fatal_by_analyst'] else 1
        fba_html += '<div class="bar-chart">'
        for analyst, count in metrics['fatal_by_analyst']:
            pct = (count / max_c * 100) if max_c else 0
            a_safe = str(analyst).replace("'", "&#39;")
            fba_html += f'''<div class="bar-row" onclick="openAnalystFatalDrill('{a_safe}')">
              <div class="bar-label">{a_safe}</div>
              <div class="bar-track"><div class="bar-fill bar-fatal" style="width:{pct}%"></div></div>
              <div class="bar-value">{count}</div>
            </div>'''
        fba_html += '</div>'
    else:
        fba_html += '<p class="muted">No fatal errors found in Analyst file.</p>'
    fba_html += '</div>'

    # ANALYST TABLE
    at_html = '''<div class="section half" id="sec-analyst">
      <div class="section-header"><h2>Analyst Performance</h2><button class="dl-btn" onclick="downloadSectionCSV('sec-analyst', 'Analyst_Performance')">Download</button></div>
      <input type="text" id="analystSearch" placeholder="Search analyst..." oninput="filterAnalystTable()" class="search-box"/>
      <div class="table-wrap"><table class="data-table" id="analystTable">
        <thead><tr>
          <th onclick="sortDataTable('analystTable',0,'str')">Analyst</th>
          <th onclick="sortDataTable('analystTable',1,'str')">Org</th>
          <th onclick="sortDataTable('analystTable',2,'num')">Audited</th>
          <th onclick="sortDataTable('analystTable',3,'num')">Accuracy</th>
          <th onclick="sortDataTable('analystTable',4,'num')">Fatal</th>
          <th onclick="sortDataTable('analystTable',5,'num')">Non-Fatal Defects</th>
        </tr></thead><tbody>'''
    for a in metrics['analysts']:
        acc_cls = rag_class(a['accuracy'])
        at_html += f'''<tr onclick="openAnalystDrill('{a['analyst']}')">
          <td>{a['analyst']}</td>
          <td>{a['org']}</td>
          <td>{a['audited']}</td>
          <td class="{acc_cls}">{fmt_pct(a['accuracy'])}</td>
          <td>{a['fatal']}</td>
          <td>{a['nonfatal']}</td>
        </tr>'''
    at_html += '</tbody></table></div></div>'

    # FILTER BAR
    def _multi_dd(label, key, values):
        opts = ''
        for v in values:
            v_safe = str(v).replace("'", "&#39;")
            opts += f'<label><input type="checkbox" value="{v_safe}" onchange="onFilterChange()"/> {v_safe}</label>'
        return f'''<div class="filter-dd" data-key="{key}">
          <button class="filter-btn" onclick="toggleDD(this)">{label} <span class="badge" id="badge-{key}">All</span> ▾</button>
          <div class="filter-panel">
            <label class="all-toggle"><input type="checkbox" checked onchange="toggleAll(this,'{key}')"/> All</label>
            <div class="filter-options">{opts}</div>
          </div>
        </div>'''
    f = metrics['filters']
    filter_html = '<div class="filter-bar">'
    filter_html += _multi_dd('Org', 'orgs', f['orgs'])
    filter_html += _multi_dd('Analyst', 'analysts', f['analysts'])
    filter_html += _multi_dd('Category', 'categories', f['categories'])
    filter_html += _multi_dd('Week', 'weeks', f['weeks'])
    filter_html += _multi_dd('Topic', 'topics', f['topics'])
    filter_html += '<button class="reset-btn" onclick="resetFilters()">Reset</button>'
    filter_html += '<span class="filter-info" id="filterInfo">Filters update KPIs and drill-downs</span></div>'

    # FOOTER
    footer = ''
    if missing_files:
        footer = f'''<div class="footer-note"><strong>Note:</strong> This process does not yet track {", ".join(missing_files)}. Cards are auto-hidden until those files are added.</div>'''

    rows_json = json.dumps(metrics['rows'], default=str)
    mom_json  = json.dumps(metrics.get('mom') or {}, default=str)
    orgs_json = json.dumps(metrics['orgs'], default=str)
    analysts_json = json.dumps(metrics['analysts'], default=str)

    # Build EMAIL_DATA - condensed data for client-side email generation
    email_data = {
        'process': process_label,
        'month': month_label,
        'dashboardUrl': '[paste your CloudFront / SharePoint dashboard URL here]',
        'headline': metrics['headline'],
        'mom': metrics.get('mom') or {},
        'goal': metrics.get('defect_reduction_goal') or {},
        'orgs': metrics['orgs'],
        'topDefects': metrics.get('top_defects', [])[:3],  # only top 3 for email
        'impact': metrics.get('impact_categorization', []),
        'status': {
            'disputes':         metrics.get('disputes'),
            'rectification':    metrics.get('rectification'),
            'ivoc':             metrics.get('ivoc'),
            'defect_reduction': metrics.get('defect_reduction'),
        },
    }
    email_data_json = json.dumps(email_data, default=str)

    html = HTML_TEMPLATE.format(
        process=process_label, month=month_label,
        filter_bar=filter_html, kpi_row=kpi_html, status_row=status_html,
        extra_row=extra_row,
        wow=wow_html, mom=mom_html, top_defects=top_html_inline, orgs=org_html, pareto=pareto_html,
        impact=impact_html,
        fatal_by_analyst=fba_html, analyst_table=at_html, footer=footer,
        rows_json=rows_json, mom_json=mom_json, orgs_json=orgs_json,
        analysts_json=analysts_json, email_data_json=email_data_json,
    )
    return html


# ============================================================
# HTML TEMPLATE
# ============================================================
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"/>
<title>QC Dashboard — {process} {month}</title>
<style>
* {{ box-sizing: border-box; }}
body {{ font-family: -apple-system, "Segoe UI", Arial, sans-serif; margin: 0; padding: 16px;
       background: #FAF8F4; color: #2C2C2A; }}
h2 {{ font-size: 14px; margin: 0 0 10px 0; color: #2C2C2A; font-weight: 500; }}
.section-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; }}
.section-header h2 {{ margin: 0; }}
.dl-btn {{ background: #F1EFE8; border: 0.5px solid #D3D1C7; color: #444441;
           padding: 4px 10px; border-radius: 6px; cursor: pointer;
           font-size: 11px; }}
.dl-btn:hover {{ background: #D3D1C7; color: #2C2C2A; }}
.row {{ display: flex; gap: 12px; margin-bottom: 16px; flex-wrap: wrap; }}
.row > * {{ flex: 1 1 0; min-width: 0; }}
.section {{ background: #FFFFFF; border-radius: 10px; padding: 14px;
            border: 0.5px solid #D3D1C7; margin-bottom: 16px; }}
.section.half {{ flex: 1 1 calc(50% - 8px); }}
.section.third {{ flex: 1 1 calc(33.333% - 11px); }}
.status-card-wrap {{ flex: 1 1 calc(50% - 8px); display: flex; gap: 12px; }}
.goal-line {{ font-size: 13px; margin: 4px 0; color: #444441; }}
.goal-met {{ background: #C0DD97; color: #173404; padding: 4px 10px;
              border-radius: 6px; font-weight: 500; font-size: 13px; }}
.goal-miss {{ background: #F7C1C1; color: #501313; padding: 4px 10px;
                border-radius: 6px; font-weight: 500; font-size: 13px; }}
.inflow-flag {{ background: #fef3c7; }}
.muted {{ color: #6b7280; font-size: 12px; }}
.muted.small {{ font-size: 11px; }}

.ph {{ font-size: 13px; color: #6b7280; margin-bottom: 8px; }}

/* Filter bar */
.filter-bar {{ background: #FFFFFF; padding: 10px; border-radius: 10px;
               display: flex; gap: 8px; flex-wrap: wrap; align-items: center;
               border: 0.5px solid #D3D1C7; margin-bottom: 16px; }}
.filter-dd {{ position: relative; }}
.filter-btn {{ background: #F1EFE8; border: 0.5px solid #D3D1C7; padding: 6px 12px;
               border-radius: 6px; cursor: pointer; font-size: 13px; color: #2C2C2A; }}
.filter-btn:hover {{ background: #D3D1C7; }}
.badge {{ background: #BA7517; color: #fff; border-radius: 10px;
          padding: 1px 6px; font-size: 11px; margin-left: 4px; }}
.filter-panel {{ display: none; position: absolute; top: 100%; left: 0;
                 background: #FFFFFF; border: 0.5px solid #D3D1C7; border-radius: 6px;
                 padding: 8px; min-width: 200px; max-height: 300px; overflow-y: auto;
                 z-index: 100; box-shadow: 0 4px 10px rgba(0,0,0,.06); }}
.filter-panel.open {{ display: block; }}
.filter-panel label {{ display: block; padding: 4px 0; font-size: 13px; cursor: pointer; }}
.all-toggle {{ font-weight: 500; border-bottom: 0.5px solid #D3D1C7; margin-bottom: 4px;
               padding-bottom: 6px !important; }}
.reset-btn {{ background: #A32D2D; color: #fff; border: none; padding: 6px 14px;
              border-radius: 6px; cursor: pointer; font-size: 13px; }}
.filter-info {{ color: #5F5E5A; font-size: 11px; margin-left: 8px; }}

/* KPI cards */
.kpi-card {{ background: #FFFFFF; padding: 16px; border-radius: 10px;
             border: 0.5px solid #D3D1C7; cursor: pointer;
             transition: transform .1s; }}
.kpi-card:hover {{ transform: translateY(-2px); border-color: #B4B2A9; }}
.kpi-label {{ font-size: 12px; color: #5F5E5A; text-transform: uppercase; letter-spacing: 0.5px; }}
.kpi-value {{ font-size: 28px; font-weight: 500; margin: 6px 0; color: #2C2C2A; }}
.kpi-sub {{ font-size: 11px; color: #5F5E5A; }}
.kpi-fatal .kpi-value {{ color: #791F1F; }}
.kpi-fatal {{ background: #FCEBEB; border-color: #F09595; }}
.kpi-fatal .kpi-label {{ color: #501313; }}
.kpi-fatal .kpi-sub {{ color: #791F1F; }}

/* RAG */
.rag-green {{ background: #EAF3DE !important; border-color: #97C459 !important; }}
.rag-amber {{ background: #FAEEDA !important; border-color: #EF9F27 !important; }}
.rag-red {{ background: #FCEBEB !important; border-color: #F09595 !important; }}
.rag-na    {{ background: #F1EFE8 !important; }}
td.rag-green {{ color: #173404; font-weight: 500; }}
td.rag-amber {{ color: #412402; font-weight: 500; }}
td.rag-red   {{ color: #501313; font-weight: 500; }}

/* Status cards - notebook style with lines between rows */
.status-card {{ background: #FFFFFF; border-radius: 10px; padding: 14px;
                border: 0.5px solid #D3D1C7; cursor: pointer;
                flex: 1 1 0; min-width: 240px; max-width: 480px; }}
.status-card:hover {{ border-color: #B4B2A9; }}
.status-header {{ font-weight: 500; font-size: 13px; color: #2C2C2A;
                  border-bottom: 0.5px solid #D3D1C7; padding-bottom: 6px; margin-bottom: 8px; }}
.status-total {{ font-size: 13px; color: #5F5E5A; margin-bottom: 8px;
                 padding-bottom: 6px; border-bottom: 0.5px solid #D3D1C7; }}
.status-row {{ display: flex; justify-content: space-between; font-size: 13px;
               padding: 6px 0; cursor: pointer;
               border-bottom: 0.5px solid #F1EFE8; }}
.status-row:last-child {{ border-bottom: none; }}
.status-row:hover {{ background: #FAF8F4; }}
.status-divider {{ height: 1px; background: #D3D1C7; margin: 6px 0; }}
.status-sla {{ margin-top: 10px; padding: 6px 8px; border-radius: 6px;
               text-align: center; font-size: 12px; font-weight: 500; }}

/* Bar chart */
.bar-chart {{ }}
.bar-row {{ display: grid; grid-template-columns: 1.6fr 3fr 50px;
            gap: 10px; align-items: center; padding: 4px 0;
            cursor: pointer; font-size: 13px; }}
.bar-row:hover {{ background: #FAF8F4; }}
.bar-label {{ overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
.bar-track {{ background: #F1EFE8; border-radius: 4px; height: 18px; overflow: hidden; }}
.bar-fill  {{ background: #BA7517; height: 100%; }}
.bar-fatal {{ background: #A32D2D; }}
.bar-value {{ text-align: right; font-weight: 500; color: #2C2C2A; }}

/* Org/Analyst tables */
.data-table {{ width: 100%; border-collapse: collapse; font-size: 13px; table-layout: auto; }}
.data-table th {{ background: #F1EFE8; padding: 8px 12px; text-align: left;
                  cursor: pointer; user-select: none; position: sticky; top: 0;
                  color: #2C2C2A; font-weight: 500; }}
.data-table td {{ padding: 6px 12px; border-bottom: 0.5px solid #F1EFE8; white-space: nowrap; }}
.data-table tr:hover {{ background: #FAF8F4; cursor: pointer; }}
.data-table th:nth-child(n+2), .data-table td:nth-child(n+2) {{ text-align: right; }}
.data-table th:nth-child(1), .data-table td:nth-child(1) {{ text-align: left; }}
.search-box {{ width: 100%; padding: 6px 10px; border: 0.5px solid #D3D1C7;
               border-radius: 6px; margin-bottom: 8px; font-size: 13px; background: #FFFFFF; }}
.table-wrap {{ max-height: 320px; overflow-y: auto; }}

/* WoW */
.wow-block {{ margin-bottom: 12px; }}
.wow-title {{ font-size: 12px; font-weight: 500; color: #5F5E5A; margin-bottom: 4px; }}
.wow-bars {{ display: flex; gap: 8px; align-items: flex-end; height: 80px; }}
.wow-bar-col {{ flex: 1; display: flex; flex-direction: column; align-items: center;
                justify-content: flex-end; height: 100%; }}
.wow-num {{ font-size: 12px; font-weight: 500; color: #2C2C2A; }}
.wow-bar {{ width: 100%; height: 60%; background: #F1EFE8; border-radius: 3px 3px 0 0;
            display: flex; flex-direction: column; justify-content: flex-end; }}
.wow-fill {{ width: 100%; border-radius: 3px 3px 0 0; transition: height .3s;
             background: #BA7517; }}
.wow-lbl {{ font-size: 11px; color: #5F5E5A; margin-top: 4px; }}

/* MoM */
.mom-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
.mom-table th {{ background: #F1EFE8; padding: 8px; text-align: center;
                 font-weight: 500; font-size: 12px; color: #2C2C2A; }}
.mom-table td {{ padding: 8px; text-align: center; }}
.mom-table .mom-lbl {{ text-align: left; font-weight: 500; color: #2C2C2A; }}
.mom-table .mom-val {{ font-size: 16px; font-weight: 700; }}
.mom-table .vs {{ color: #9ca3af; font-size: 11px; }}
.mom-table .mom-delta {{ font-size: 11px; font-weight: 500; margin-top: 2px; }}
.mom-table .mom-fatal {{ cursor: pointer; color: #dc2626; }}
.arr-up {{ color: #b91c1c; font-weight: 700; }}
.arr-dn {{ color: #047857; font-weight: 700; }}

/* Modal */
.modal-overlay {{ position: fixed; inset: 0; background: rgba(0,0,0,.5);
                  display: none; align-items: center; justify-content: center; z-index: 200; }}
.modal-overlay.open {{ display: flex; }}
.modal {{ background: #fff; border-radius: 10px; padding: 20px;
          max-width: 90%; max-height: 85vh; overflow-y: auto; min-width: 600px; }}
.modal-header {{ display: flex; justify-content: space-between; align-items: center;
                 border-bottom: 1px solid #e5e7eb; padding-bottom: 10px; margin-bottom: 12px; }}
.modal-header h3 {{ margin: 0; font-size: 16px; }}
.modal-close {{ background: #ef4444; color: #fff; border: none; padding: 4px 10px;
                border-radius: 6px; cursor: pointer; }}
.modal-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
.modal-table th {{ background: #f3f4f6; padding: 6px 8px; text-align: left; }}
.modal-table td {{ padding: 6px 8px; border-bottom: 1px solid #f3f4f6; vertical-align: top; }}

.footer-note {{ background: #fef3c7; border-left: 4px solid #f59e0b;
                padding: 10px 14px; border-radius: 6px; font-size: 12px;
                color: #92400e; margin-top: 16px; }}
</style></head>
<body>
<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;margin-bottom:16px;">
  <div class="ph" style="margin:0">{process} • {month}</div>
  <div style="display:flex;gap:8px;">
    <button onclick="sendEmail()" style="background:#BA7517;color:white;border:none;padding:10px 18px;border-radius:6px;cursor:pointer;font-size:14px;font-weight:500;">📧 Send Email</button>
  </div>
</div>

{filter_bar}

{kpi_row}

{extra_row}

<div class="row">
  {wow}
  {mom}
</div>

{status_row}

<div class="row">
  {top_defects}
  {impact}
</div>

<div class="row">
  {orgs}
  {pareto}
</div>

<div class="row">
  {fatal_by_analyst}
  {analyst_table}
</div>

{footer}

<div id="modalOverlay" class="modal-overlay" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <div class="modal-header"><h3 id="modalTitle">Drill-down</h3>
      <button class="modal-close" onclick="closeModal()">Close</button></div>
    <div id="modalBody"></div>
  </div>
</div>

<script>
const ROWS = {rows_json};
const MOM = {mom_json};
const ORGS = {orgs_json};
const ANALYSTS = {analysts_json};
const EMAIL_DATA = {email_data_json};
let activeFilters = {{ orgs: [], analysts: [], categories: [], weeks: [], topics: [] }};

// CSV Download Helper - exports an HTML element's table or list as CSV
function downloadSectionCSV(sectionId, filename) {{
  const section = document.getElementById(sectionId);
  if (!section) return;
  const rows = [];

  // Try table first
  const table = section.querySelector('table');
  if (table) {{
    const trs = table.querySelectorAll('tr');
    trs.forEach(tr => {{
      const cells = tr.querySelectorAll('th, td');
      const rowData = Array.from(cells).map(c => {{
        let txt = c.innerText.replace(/"/g, '""').replace(/\\n/g, ' ').trim();
        return '"' + txt + '"';
      }});
      if (rowData.length > 0) rows.push(rowData.join(','));
    }});
  }}

  // Try bar-chart rows
  if (rows.length === 0) {{
    const bars = section.querySelectorAll('.bar-row');
    if (bars.length > 0) {{
      rows.push('"Item","Value"');
      bars.forEach(b => {{
        const lbl = b.querySelector('.bar-label')?.innerText || '';
        const val = b.querySelector('.bar-value')?.innerText || '';
        rows.push(`"${{lbl.replace(/"/g, '""')}}","${{val}}"`);
      }});
    }}
  }}

  // Try status-row breakdowns
  if (rows.length === 0) {{
    const srows = section.querySelectorAll('.status-row');
    if (srows.length > 0) {{
      rows.push('"Category","Count"');
      srows.forEach(r => {{
        const spans = r.querySelectorAll('span, strong');
        if (spans.length >= 2) {{
          rows.push(`"${{spans[0].innerText.replace(/"/g, '""')}}","${{spans[1].innerText}}"`);
        }}
      }});
    }}
  }}

  if (rows.length === 0) {{
    alert('No data to export from this section.');
    return;
  }}

  // Use char codes to avoid Python f-string escape ambiguity
  const NL = String.fromCharCode(13) + String.fromCharCode(10);
  const BOM = String.fromCharCode(0xFEFF);
  const csv = rows.join(NL);
  const blob = new Blob([BOM + csv], {{ type: 'text/csv;charset=utf-8;' }});
  const url = URL.createObjectURL(blob);
  const a = document.createElement('a');
  a.href = url;
  a.download = filename + '.csv';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
  URL.revokeObjectURL(url);
}}

// ============================================================
// EMAIL BODY GENERATION
// ============================================================
function buildEmailBodyHTML() {{
  // Build a static email body from current dashboard data.
  // All inline styles, table-based, no JS - Outlook-compatible.
  const h = EMAIL_DATA.headline;
  const mom = EMAIL_DATA.mom || {{}};
  const goal = EMAIL_DATA.goal || {{}};
  const orgs = EMAIL_DATA.orgs || [];
  const topDefects = EMAIL_DATA.topDefects || [];
  const impact = EMAIL_DATA.impact || [];
  const status = EMAIL_DATA.status || {{}};
  const process = EMAIL_DATA.process;
  const month = EMAIL_DATA.month;
  const dashUrl = EMAIL_DATA.dashboardUrl;

  // Use bgcolor attribute (HTML4) + inline styles for Outlook dark-mode resistance
  const TD = 'bgcolor="#FFFFFF" style="padding:8px 12px;border:1px solid #D3D1C7;font-family:Arial,sans-serif;font-size:13px;color:#2C2C2A;background-color:#FFFFFF;"';
  const TDR = 'bgcolor="#FFFFFF" style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;font-family:Arial,sans-serif;font-size:13px;color:#2C2C2A;background-color:#FFFFFF;"';
  const TH = 'bgcolor="#F1EFE8" style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;background-color:#F1EFE8;font-weight:500;text-align:left;font-family:Arial,sans-serif;font-size:13px;color:#2C2C2A;"';
  const THR = 'bgcolor="#F1EFE8" style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;background-color:#F1EFE8;font-weight:500;text-align:right;font-family:Arial,sans-serif;font-size:13px;color:#2C2C2A;"';

  let html = '<div style="font-family:Arial,sans-serif;color:#2C2C2A;max-width:760px;font-size:14px;line-height:1.5;">';

  html += '<p style="color:#791F1F;font-style:italic;">&lt;Please do not forward&gt;</p>';
  html += '<p>Dear Team,</p>';
  html += `<p>Please find the <strong>${{month}} QC Audit Metrics for ${{process}}</strong> below.</p>`;
  html += '<p>📄 Refer to <em>Page 0</em> for definitions and methodology: <strong>[your link]</strong></p>';

  html += '<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>';

  // EXECUTIVE SUMMARY PLACEHOLDER
  html += '<h3 style="color:#2C2C2A;font-weight:500;">Executive Summary</h3>';
  html += '<ul>';
  html += '<li>[Add commentary on this month performance, key changes, watch items]</li>';
  html += '<li>[Mention top analysts with fatal errors here]</li>';
  html += '<li>[Recommended actions for May]</li>';
  html += '</ul>';

  html += '<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>';

  // KPI TABLE
  html += `<h3 style="color:#2C2C2A;font-weight:500;">Key Metrics — ${{month}}</h3>`;
  html += '<table style="border-collapse:collapse;width:auto;margin:0 0 20px 0;"><tbody>';
  html += `<tr><td ${{TH}}>Cases Audited</td><td ${{TDR}}>${{h.audited}}</td></tr>`;
  let accStr = h.accuracy.toFixed(2) + '%';
  if (mom.cur_anl_acc != null) accStr += ' (Process) / ' + mom.cur_anl_acc.toFixed(2) + '% (Analyst)';
  html += `<tr><td ${{TH}}>Accuracy</td><td ${{TDR}}>${{accStr}}</td></tr>`;
  let fatalStr = `${{h.fatal}} (Inap ${{h.inap}}, Conf ${{h.conf}})`;
  if (h.analyst_fatal != null) fatalStr += ` / Analyst: ${{h.analyst_fatal}}`;
  html += `<tr><td ${{TH}}>Fatal Errors</td><td ${{TDR}}>${{fatalStr}}</td></tr>`;
  html += `<tr><td ${{TH}}>Total Defects</td><td ${{TDR}}>${{h.defects}}</td></tr>`;
  html += '</tbody></table>';

  // DEFECT REDUCTION GOAL BANNER
  if (goal && goal.actual_change_pct != null) {{
    const badge = goal.goal_met ? '✅ MET' : '❌ NOT MET';
    const bg = goal.goal_met ? '#EAF3DE' : '#FCEBEB';
    const fg = goal.goal_met ? '#173404' : '#501313';
    html += `<div style="background:${{bg}};border:1px solid ${{fg}};padding:12px;border-radius:6px;margin:0 0 20px 0;color:${{fg}};">`;
    html += `<strong>Defect Reduction Goal: ${{badge}}</strong> — Target −${{goal.target_pct.toFixed(0)}}% MoM (Inappropriate Resolution). `;
    html += `${{goal.prev_label}}: ${{goal.prev_count}} → ${{goal.cur_label}}: ${{goal.cur_count}}. Actual: ${{goal.actual_change_pct.toFixed(1)}}%.`;
    html += '</div>';
  }}

  // MoM COMPARISON
  if (mom && mom.prev_label) {{
    html += '<h3 style="color:#2C2C2A;font-weight:500;">Month-over-Month Comparison</h3>';
    html += '<table style="border-collapse:collapse;width:100%;margin:0 0 20px 0;"><thead><tr>';
    html += `<th ${{TH}}>Metric</th><th ${{THR}}>${{mom.prev_label}}</th><th ${{THR}}>${{mom.cur_label}}</th><th ${{THR}}>Δ</th>`;
    html += '</tr></thead><tbody>';
    html += `<tr><td ${{TD}}>Cases Audited</td><td ${{TDR}}>${{mom.prev_audited||0}}</td><td ${{TDR}}>${{mom.cur_audited||0}}</td><td ${{TDR}}>${{(mom.audited_delta>0?'+':'')+(mom.audited_delta||0)}}</td></tr>`;
    html += `<tr><td ${{TD}}>Process Accuracy</td><td ${{TDR}}>${{(mom.prev_proc_acc||0).toFixed(2)}}%</td><td ${{TDR}}>${{(mom.cur_proc_acc||0).toFixed(2)}}%</td><td ${{TDR}}>${{(mom.proc_acc_delta>0?'+':'')+(mom.proc_acc_delta||0).toFixed(2)}}</td></tr>`;
    if (mom.cur_anl_acc != null) html += `<tr><td ${{TD}}>Analyst Accuracy</td><td ${{TDR}}>${{(mom.prev_anl_acc||0).toFixed(2)}}%</td><td ${{TDR}}>${{(mom.cur_anl_acc||0).toFixed(2)}}%</td><td ${{TDR}}>${{(mom.anl_acc_delta>0?'+':'')+(mom.anl_acc_delta||0).toFixed(2)}}</td></tr>`;
    html += `<tr><td ${{TD}}>Fatal Errors</td><td ${{TDR}}>${{mom.prev_fatal||0}}</td><td ${{TDR}}>${{mom.cur_fatal||0}}</td><td ${{TDR}}>${{(mom.fatal_delta>0?'+':'')+(mom.fatal_delta||0)}}</td></tr>`;
    html += '</tbody></table>';
  }}

  // STATUS SUMMARY (4 cards as table rows)
  html += '<h3 style="color:#2C2C2A;font-weight:500;">Status Summary</h3>';
  html += '<table style="border-collapse:collapse;width:100%;margin:0 0 20px 0;"><thead><tr>';
  html += `<th ${{TH}}>Status Type</th><th ${{THR}}>Total</th><th ${{THR}}>SLA %</th>`;
  html += '</tr></thead><tbody>';
  const statusKeys = [['disputes','Disputes'],['rectification','Rectification'],['ivoc','IVOC'],['defect_reduction','Defect Reduction']];
  for (const [k, lbl] of statusKeys) {{
    const s = status[k];
    if (s && s.total != null) {{
      const slaStr = s.sla_total > 0 ? `${{s.sla_met}}/${{s.sla_total}} (${{s.sla_pct.toFixed(1)}}%)` : 'N/A';
      html += `<tr><td ${{TD}}>${{lbl}}</td><td ${{TDR}}>${{s.total}}</td><td ${{TDR}}>${{slaStr}}</td></tr>`;
    }}
  }}
  html += '</tbody></table>';

  // ORG SCORECARD
  if (orgs && orgs.length > 0) {{
    html += '<h3 style="color:#2C2C2A;font-weight:500;">Org Scorecard</h3>';
    html += '<table style="border-collapse:collapse;width:100%;margin:0 0 20px 0;"><thead><tr>';
    html += `<th ${{TH}}>Org</th><th ${{THR}}>Cases Audited</th><th ${{THR}}>Accuracy</th><th ${{THR}}>Fatal</th><th ${{THR}}>Non-Fatal Defects</th>`;
    html += '</tr></thead><tbody>';
    for (const o of orgs) {{
      html += `<tr><td ${{TD}}>${{o.org}}</td><td ${{TDR}}>${{o.audited}}</td><td ${{TDR}}>${{o.accuracy.toFixed(2)}}%</td><td ${{TDR}}>${{o.fatal}}</td><td ${{TDR}}>${{o.nonfatal}}</td></tr>`;
    }}
    html += '</tbody></table>';
  }}

  // TOP 3 DEFECTS
  if (topDefects && topDefects.length > 0) {{
    html += '<h3 style="color:#2C2C2A;font-weight:500;">Top 3 Defects — Drivers and Potential Causes</h3>';
    html += '<table style="border-collapse:collapse;width:100%;margin:0 0 20px 0;"><thead><tr>';
    html += `<th ${{TH}}>Parameter</th><th ${{THR}}>Defects</th><th ${{THR}}>Contribution %</th><th ${{TH}}>Potential Causes</th>`;
    html += '</tr></thead><tbody>';
    const totalDefects = h.defects || 1;
    for (const [name, cnt] of topDefects.slice(0, 3)) {{
      const pct = (cnt / totalDefects * 100).toFixed(1);
      html += `<tr><td ${{TD}}>${{name}}</td><td ${{TDR}}>${{cnt}}</td><td ${{TDR}}>${{pct}}%</td><td ${{TD}} style="color:#5F5E5A;font-style:italic;">[Fill potential causes]</td></tr>`;
    }}
    html += '</tbody></table>';
  }}

  // IMPACT CATEGORIZATION
  if (impact && impact.length > 0) {{
    html += '<h3 style="color:#2C2C2A;font-weight:500;">Defect Impact Categorization</h3>';
    html += '<table style="border-collapse:collapse;width:100%;margin:0 0 20px 0;"><thead><tr>';
    html += `<th ${{TH}}>Category</th><th ${{THR}}>Defects</th>`;
    html += '</tr></thead><tbody>';
    for (const r of impact) {{
      html += `<tr><td ${{TD}}>${{r.category}}</td><td ${{TDR}}>${{r.defects}}</td></tr>`;
    }}
    html += '</tbody></table>';
  }}

  // TREND PLACEHOLDER
  html += '<h3 style="color:#2C2C2A;font-weight:500;">Trend Analysis (Multi-Month)</h3>';
  html += '<p style="color:#5F5E5A;font-style:italic;">[Insert multi-month trending table from past audit cycles. Manual entry.]</p>';

  // DASHBOARD LINK
  html += '<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>';
  html += '<h3 style="color:#2C2C2A;font-weight:500;">Interactive QC Dashboard</h3>';
  html += `<p><strong>🔗 Access:</strong> <a href="${{dashUrl}}" style="color:#BA7517;">${{dashUrl}}</a></p>`;
  html += '<p>Dashboard capabilities:</p><ul>';
  html += '<li>Live filtering across org, analyst, week, audit category, and defect topic</li>';
  html += '<li>Drill-down on any KPI, chart, or table row for case-level audit commentary</li>';
  html += '<li>Section-level CSV export for ad-hoc analysis</li>';
  html += '<li>SLA tracking for Disputes, Rectification, IVOC, and Defect Reduction</li>';
  html += '<li>Defect Impact Categorization by business impact</li>';
  html += '<li>Defect Reduction Goal tracker against 10% MoM target</li>';
  html += '</ul>';

  // CALIBRATION
  html += '<h3 style="color:#2C2C2A;font-weight:500;">Calibration Cycle</h3>';
  html += '<table style="border-collapse:collapse;width:100%;margin:0 0 20px 0;"><thead><tr>';
  html += `<th ${{TH}}>Calibration with Ops Team</th><th ${{TH}}>Parameter Revision</th><th ${{TH}}>QC SPOC</th><th ${{TH}}>QC Manager</th>`;
  html += '</tr></thead><tbody><tr>';
  html += `<td ${{TD}} style="color:#5F5E5A;">[Last / Next]</td><td ${{TD}} style="color:#5F5E5A;">[Last / Next]</td><td ${{TD}} style="color:#5F5E5A;">[SPOC names]</td><td ${{TD}} style="color:#5F5E5A;">[Manager LDAP]</td>`;
  html += '</tr></tbody></table>';

  html += '<p>Regards,<br/>Niki<br/>QC | Retail FinCoM<br/>Amazon</p>';
  html += '</div>';

  return html;
}}

function sendEmail() {{
  const html = buildEmailBodyHTML();
  const subject = `APQC_Retail FinCoM ${{EMAIL_DATA.process}} All-Orgs Audit Metrics_${{EMAIL_DATA.month}}`;

  // Step 1: Try to copy HTML to clipboard
  let clipboardOk = false;

  const tryClipboard = async () => {{
    try {{
      if (navigator.clipboard && window.ClipboardItem) {{
        const htmlBlob = new Blob([html], {{ type: 'text/html' }});
        const textBlob = new Blob([html.replace(/<[^>]*>/g, '')], {{ type: 'text/plain' }});
        const item = new ClipboardItem({{ 'text/html': htmlBlob, 'text/plain': textBlob }});
        await navigator.clipboard.write([item]);
        return true;
      }} else if (navigator.clipboard && navigator.clipboard.writeText) {{
        await navigator.clipboard.writeText(html);
        return true;
      }}
    }} catch (err) {{
      console.warn('Clipboard write failed:', err);
    }}
    return false;
  }};

  tryClipboard().then(ok => {{
    clipboardOk = ok;
    if (ok) {{
      // Success path: use hidden anchor with target=_blank so Outlook opens
      // via OS handoff without navigating the dashboard tab away
      const mailto = `mailto:?subject=${{encodeURIComponent(subject)}}`;
      const a = document.createElement('a');
      a.href = mailto;
      a.target = '_blank';
      a.rel = 'noopener noreferrer';
      document.body.appendChild(a);
      a.click();
      document.body.removeChild(a);
      // Show confirmation
      setTimeout(() => {{
        alert('✅ Email body copied to clipboard.\\n\\nOutlook is opening — click in the body area and press Ctrl+V to paste.\\n\\nThen:\\n• Add Executive Summary at the top\\n• Update Page 0 link\\n• Update Calibration cycle\\n• Review and Send');
      }}, 100);
    }} else {{
      // Fallback: show modal with email body inside a textarea for manual copy
      showEmailFallback(html, subject);
    }}
  }});
}}

function showEmailFallback(html, subject) {{
  // Create modal overlay
  const overlay = document.createElement('div');
  overlay.style.cssText = 'position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.6);z-index:9999;display:flex;align-items:center;justify-content:center;padding:20px;';

  const modal = document.createElement('div');
  modal.style.cssText = 'background:#FFFFFF;border-radius:10px;padding:24px;max-width:800px;width:100%;max-height:90vh;overflow-y:auto;box-shadow:0 10px 40px rgba(0,0,0,0.3);';

  modal.innerHTML = `
    <h2 style="margin:0 0 12px 0;color:#2C2C2A;">📧 Email Body Ready</h2>
    <p style="color:#5F5E5A;font-size:13px;margin:0 0 16px 0;">Clipboard access was blocked. Use the steps below instead:</p>
    <ol style="font-size:13px;color:#2C2C2A;line-height:1.7;">
      <li>Click the <strong>Select All</strong> button below</li>
      <li>Press <strong>Ctrl+C</strong> to copy</li>
      <li>Click <strong>Open in Outlook</strong> — Outlook opens with subject pre-filled</li>
      <li>Click in the body area, press <strong>Ctrl+V</strong> to paste</li>
      <li>Add Executive Summary, Page 0 link, Calibration cycle — review — Send</li>
    </ol>
    <div style="display:flex;gap:8px;margin:12px 0;">
      <button id="selectAllBtn" style="background:#BA7517;color:white;border:none;padding:8px 14px;border-radius:6px;cursor:pointer;font-size:13px;">📋 Select All</button>
      <button id="openOutlookBtn" style="background:#2C2C2A;color:white;border:none;padding:8px 14px;border-radius:6px;cursor:pointer;font-size:13px;">📧 Open in Outlook</button>
      <button id="closeModalBtn" style="background:#F1EFE8;color:#2C2C2A;border:0.5px solid #D3D1C7;padding:8px 14px;border-radius:6px;cursor:pointer;font-size:13px;margin-left:auto;">Close</button>
    </div>
    <div id="emailContent" contenteditable="true" style="border:0.5px solid #D3D1C7;border-radius:6px;padding:16px;max-height:400px;overflow-y:auto;background:#FAF8F4;font-size:12px;">${{html}}</div>
  `;

  overlay.appendChild(modal);
  document.body.appendChild(overlay);

  document.getElementById('selectAllBtn').onclick = () => {{
    const content = document.getElementById('emailContent');
    const range = document.createRange();
    range.selectNodeContents(content);
    const sel = window.getSelection();
    sel.removeAllRanges();
    sel.addRange(range);
  }};
  document.getElementById('openOutlookBtn').onclick = () => {{
    window.location.href = `mailto:?subject=${{encodeURIComponent(subject)}}`;
  }};
  document.getElementById('closeModalBtn').onclick = () => {{
    document.body.removeChild(overlay);
  }};
  overlay.onclick = (e) => {{
    if (e.target === overlay) document.body.removeChild(overlay);
  }};
}}


// Filter dropdown UI
function toggleDD(btn) {{
  document.querySelectorAll('.filter-panel.open').forEach(p => {{
    if (p !== btn.nextElementSibling) p.classList.remove('open');
  }});
  btn.nextElementSibling.classList.toggle('open');
}}
document.addEventListener('click', e => {{
  if (!e.target.closest('.filter-dd')) {{
    document.querySelectorAll('.filter-panel.open').forEach(p => p.classList.remove('open'));
  }}
}});
function toggleAll(checkbox, key) {{
  const panel = checkbox.closest('.filter-panel');
  panel.querySelectorAll('.filter-options input[type=checkbox]').forEach(b => b.checked = false);
  onFilterChange();
}}
function onFilterChange() {{
  ['orgs','analysts','categories','weeks','topics'].forEach(k => {{
    const dd = document.querySelector(`[data-key='${{k}}']`);
    const checked = Array.from(dd.querySelectorAll('.filter-options input:checked')).map(b => b.value);
    activeFilters[k] = checked;
    document.getElementById('badge-' + k).textContent = checked.length === 0 ? 'All' : `${{checked.length}}`;
  }});
  filterAnalystTable();
  recomputeKPIs();
}}

function recomputeKPIs() {{
  // Recompute the 4 headline KPIs from filtered rows
  const filtered = applyFilters(ROWS);
  const total = filtered.length;

  // Audited
  document.getElementById('kpi-audited').textContent = total;

  // Accuracy (average)
  let acc = '—';
  let accClass = 'rag-na';
  if (total > 0) {{
    const accSum = filtered.reduce((s, r) => s + (r.accuracy || 0), 0);
    const accAvg = accSum / total;
    acc = accAvg.toFixed(1) + '%';
    if (accAvg >= 95) accClass = 'rag-green';
    else if (accAvg >= 90) accClass = 'rag-amber';
    else accClass = 'rag-red';
  }}
  document.getElementById('kpi-accuracy').textContent = acc;
  const accCard = document.getElementById('kpi-accuracy-card');
  accCard.classList.remove('rag-green','rag-amber','rag-red','rag-na');
  accCard.classList.add(accClass);

  // Fatal: count param hits for Inap and Conf
  let inap = 0, conf = 0;
  filtered.forEach(r => {{
    r.params.forEach(p => {{
      const np = p.toLowerCase().replace(/[^a-z]/g,'');
      if (np.includes('appropriateresolution')) inap++;
      if (np.includes('confidentiality')) conf++;
    }});
  }});
  const fatal = inap + conf;
  document.getElementById('kpi-fatal').textContent = fatal;
  document.getElementById('kpi-fatal-sub').innerHTML =
    `Inappropriate Resolution: ${{inap}} &nbsp;|&nbsp; Confidentiality: ${{conf}}`;

  // Defects: sum of missed parameters
  const defects = filtered.reduce((s, r) => s + (r.missed || 0), 0);
  const defectRate = total > 0 ? (defects / total * 100) : 0;
  document.getElementById('kpi-defects').textContent = defects;
  document.getElementById('kpi-defects-sub').textContent = 'Total missed parameters';
  const defCard = document.getElementById('kpi-defects-card');
  defCard.classList.remove('rag-green','rag-amber','rag-red','rag-na');
  if (defectRate <= 5) defCard.classList.add('rag-green');
  else if (defectRate <= 10) defCard.classList.add('rag-amber');
  else defCard.classList.add('rag-red');
}}
function resetFilters() {{
  document.querySelectorAll('.filter-options input:checked').forEach(b => b.checked = false);
  document.querySelectorAll('.all-toggle input').forEach(b => b.checked = true);
  onFilterChange();
}}

function applyFilters(rows) {{
  return rows.filter(r => {{
    if (activeFilters.orgs.length && !activeFilters.orgs.includes(r.org)) return false;
    if (activeFilters.analysts.length && !activeFilters.analysts.includes(r.analyst)) return false;
    if (activeFilters.categories.length && !activeFilters.categories.includes(r.category)) return false;
    if (activeFilters.weeks.length && !activeFilters.weeks.includes(r.week)) return false;
    if (activeFilters.topics.length && !activeFilters.topics.includes(r.topic)) return false;
    return true;
  }});
}}

function buildTable(headers, rows) {{
  if (rows.length === 0) return '<p style="color:#6b7280">No matching cases.</p>';
  let html = '<table class="modal-table"><thead><tr>';
  headers.forEach(h => html += `<th>${{h}}</th>`);
  html += '</tr></thead><tbody>';
  rows.forEach(r => {{
    html += '<tr>';
    r.forEach(c => html += `<td>${{c}}</td>`);
    html += '</tr>';
  }});
  return html + '</tbody></table>';
}}

function openModal(title, html) {{
  document.getElementById('modalTitle').textContent = title;
  document.getElementById('modalBody').innerHTML = html;
  document.getElementById('modalOverlay').classList.add('open');
}}
function closeModal() {{ document.getElementById('modalOverlay').classList.remove('open'); }}

// Table sorting
function sortDataTable(tableId, idx, type) {{
  const tbl = document.getElementById(tableId);
  const tbody = tbl.querySelector('tbody');
  const dir = tbl.dataset.sortDir === 'desc' && tbl.dataset.sortIdx == idx ? 'asc' : 'desc';
  tbl.dataset.sortDir = dir;
  tbl.dataset.sortIdx = idx;
  const rows = Array.from(tbody.querySelectorAll('tr'));
  rows.sort((a, b) => {{
    let av = a.children[idx].textContent.trim();
    let bv = b.children[idx].textContent.trim();
    if (type === 'num') {{
      av = parseFloat(av.replace('%','').replace(',','')) || 0;
      bv = parseFloat(bv.replace('%','').replace(',','')) || 0;
      return dir === 'desc' ? bv - av : av - bv;
    }}
    return dir === 'desc' ? bv.localeCompare(av) : av.localeCompare(bv);
  }});
  rows.forEach(r => tbody.appendChild(r));
}}

function filterAnalystTable() {{
  const q = (document.getElementById('analystSearch').value || '').toLowerCase();
  const tbl = document.getElementById('analystTable');
  for (const tr of tbl.querySelectorAll('tbody tr')) {{
    const cells = tr.children;
    const name = cells[0].textContent.toLowerCase();
    const org = cells[1].textContent;
    let show = true;
    if (q && !name.includes(q)) show = false;
    if (activeFilters.orgs.length && !activeFilters.orgs.includes(org)) show = false;
    tr.style.display = show ? '' : 'none';
  }}
}}

// ----- Drill-downs -----
function openDrill(kind) {{
  const filtered = applyFilters(ROWS);
  let title='', headers=[], rows=[];
  if (kind === 'audited') {{
    title = `Cases Audited (${{filtered.length}})`;
    headers = ['Case','Analyst','Org','Date','Accuracy'];
    rows = filtered.map(r => [r.case, r.analyst, r.org, r.date, r.accuracy + '%']);
  }} else if (kind === 'accuracy') {{
    title = 'Accuracy by Case (lowest first)';
    headers = ['Case','Analyst','Org','Accuracy','Defects'];
    const sorted = [...filtered].sort((a,b) => a.accuracy - b.accuracy);
    rows = sorted.map(r => [r.case, r.analyst, r.org, r.accuracy + '%', r.params.join(', ')]);
  }} else if (kind === 'fatal') {{
    // One row per fatal mark — case with both fatal params hits shows TWICE
    // so total rows match KPI's simple sum (Inap + Conf).
    const fatalRows = [];
    filtered.forEach(r => {{
      const hasInap = r.params.some(p => p.toLowerCase().replace(/[^a-z]/g,'').includes('appropriateresolution'));
      const hasConf = r.params.some(p => p.toLowerCase().replace(/[^a-z]/g,'').includes('confidentiality'));
      if (hasInap) fatalRows.push([r.case, r.analyst, r.comment || '(no comment)']);
      if (hasConf) fatalRows.push([r.case, r.analyst, r.comment || '(no comment)']);
    }});
    title = `Fatal Errors (${{fatalRows.length}})`;
    headers = ['Case Number','Analyst','Audit Comment'];
    rows = fatalRows;
  }} else if (kind === 'defects') {{
    const total = filtered.reduce((s,r) => s + r.missed, 0);
    title = `Defects — Total Missed Parameters: ${{total}}`;
    headers = ['Case','Analyst','Org','Missed','Parameters','Comment'];
    rows = filtered.filter(r => r.missed > 0).map(r => [r.case, r.analyst, r.org, r.missed, r.params.join(', '), r.comment]);
  }}
  openModal(title, buildTable(headers, rows));
}}

function openParamDrill(param) {{
  const filtered = applyFilters(ROWS).filter(r => r.params.includes(param));
  openModal(`Cases hitting: ${{param}} (${{filtered.length}})`,
            buildTable(['Case','Analyst','Org','Comment'],
                       filtered.map(r => [r.case, r.analyst, r.org, r.comment])));
}}

function openOrgDrill(org) {{
  const filtered = applyFilters(ROWS).filter(r => r.org === org);
  openModal(`${{org}} — Audits (${{filtered.length}})`,
            buildTable(['Case','Analyst','Manager','Date','Accuracy','Fatal','Defects','Comment'],
                       filtered.map(r => [r.case, r.analyst, r.manager || '(none)', r.date, r.accuracy + '%',
                                          r.is_fatal ? 'Yes' : '', r.missed, r.comment])));
}}

function openAnalystFatalDrill(analyst) {{
  // One row per fatal mark (matches sum-based fatal count)
  const filtered = applyFilters(ROWS).filter(r => r.analyst === analyst);
  const fatalRows = [];
  filtered.forEach(r => {{
    const hasInap = r.params.some(p => p.toLowerCase().replace(/[^a-z]/g,'').includes('appropriateresolution'));
    const hasConf = r.params.some(p => p.toLowerCase().replace(/[^a-z]/g,'').includes('confidentiality'));
    if (hasInap) fatalRows.push([r.case, r.comment || '(no comment)']);
    if (hasConf) fatalRows.push([r.case, r.comment || '(no comment)']);
  }});
  openModal(`${{analyst}} — Fatal Errors (${{fatalRows.length}})`,
            buildTable(['Case Number','Audit Comment'], fatalRows));
}}

function openAnalystDrill(analyst) {{
  const filtered = applyFilters(ROWS).filter(r => r.analyst === analyst);
  openModal(`${{analyst}} — Audits (${{filtered.length}})`,
            buildTable(['Case','Org','Date','Accuracy','Defects','Comment'],
                       filtered.map(r => [r.case, r.org, r.date, r.accuracy + '%', r.params.join(', '), r.comment])));
}}

function openStatusDrill(card, label) {{
  openModal(`${{card}} — ${{label}}`,
            '<p style="color:#6b7280">Detailed table for this status segment will be available in a future iteration.</p>');
}}

function openMoMDrill(kind) {{
  if (!MOM || Object.keys(MOM).length === 0) return;
  // Org-only breakdown for current month
  let html = '';
  if (kind === 'fatal') {{
    const fatalRows = ROWS.filter(r => r.is_fatal);
    const buckets = {{}};
    fatalRows.forEach(r => {{
      const key = r.org || 'Unknown';
      buckets[key] = (buckets[key] || 0) + 1;
    }});
    const sorted = Object.entries(buckets).sort((a,b) => b[1] - a[1]);
    html = buildTable(['Org', 'Fatal Count'], sorted.map(([k,v]) => [k, v]));
    openModal(`Fatal Errors — ${{MOM.cur_label}} (by Org)`, html);
  }} else if (kind === 'proc_accuracy' || kind === 'anl_accuracy') {{
    // Bottom 10 by accuracy (works for both Process and Analyst drill, since
    // ANALYSTS table aggregates per analyst from the Analyst file source)
    const sorted = [...ANALYSTS].sort((a,b) => a.accuracy - b.accuracy).slice(0, 10);
    html = buildTable(['Analyst','Org','Audited','Accuracy','Fatal','Non-Fatal Defects'],
                      sorted.map(a => [a.analyst, a.org, a.audited, a.accuracy + '%', a.fatal, a.nonfatal]));
    const label = kind === 'proc_accuracy' ? 'Process' : 'Analyst';
    openModal(`${{label}} Accuracy — Bottom 10 (${{MOM.cur_label}})`, html);
  }}
}}
</script>
</body></html>
"""


# ============================================================
# EXCEL REPORT BUILDER
# ============================================================
def build_excel(metrics, process_label, month_label, prev_month_label, out_path,
                disputes_df, rect_df, ivoc_df, defred_df):
    """Build a 6-sheet Excel report.
    Sheets: Summary | Performance (Org+Analyst+Defects) | Disputes | Rectification | IVOC | Defect Reduction"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  WARNING: openpyxl not installed; skipping Excel report.")
        return None

    wb = openpyxl.Workbook()
    thin = Side(border_style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='E5E7EB')
    hdr_font = Font(bold=True, color='1F2937')
    title_font = Font(bold=True, size=14, color='1F2937')
    section_font = Font(bold=True, size=12, color='1F2937')
    green = PatternFill('solid', fgColor='EAF3DE')
    amber = PatternFill('solid', fgColor='FAEEDA')
    red   = PatternFill('solid', fgColor='FCEBEB')

    def color_for(pct, inverted=False, sla=False):
        if pct is None: return None
        try: pct = float(pct)
        except: return None
        if sla: return green if pct >= 100 else red
        if inverted:
            if pct <= 5: return green
            if pct <= 10: return amber
            return red
        if pct >= 95: return green
        if pct >= 90: return amber
        return red

    def write_table(ws, headers, rows, start_row=1, color_cols=None, freeze=True):
        """Write a table with headers + data. color_cols = {col_idx: ('rag', inverted, sla)}"""
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = border
        for r_idx, row in enumerate(rows, start_row + 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if color_cols and c in color_cols:
                    rule = color_cols[c]
                    if rule[0] == 'rag':
                        try:
                            num = float(str(val).replace('%','').replace(',','').strip())
                            f = color_for(num, inverted=rule[1], sla=rule[2])
                            if f: cell.fill = f
                        except Exception:
                            pass
        # Auto-size columns
        for c in range(1, len(headers) + 1):
            try:
                max_len = max(len(str(headers[c-1])),
                              max((len(str(r[c-1])) if c-1 < len(r) else 0 for r in rows), default=10))
            except Exception:
                max_len = 15
            ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 4, 12), 60)
        if freeze:
            ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        # Excel auto-filter
        if rows:
            last_col = get_column_letter(len(headers))
            last_row = start_row + len(rows)
            ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"

    h = metrics['headline']

    # ============ Sheet 1: Summary ============
    ws = wb.active
    ws.title = "Summary"
    ws.cell(1, 1, f"QC Report — {process_label} {month_label}").font = title_font
    ws.cell(2, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color='6B7280')
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # KPIs
    ws.cell(4, 1, "Headline KPIs").font = section_font
    kpi_rows = [
        ['Cases Audited',                    h['audited']],
        ['Accuracy',                         f"{h['accuracy']:.2f}%"],
        ['Fatal Errors (Process)',           h['fatal']],
        ['  Inappropriate Resolution',       h['inap']],
        ['  Confidentiality',                h['conf']],
        ['Fatal Errors (Analyst)',           h['analyst_fatal'] if h['analyst_fatal'] is not None else 'NA'],
        ['  Inappropriate Resolution',       h['analyst_inap'] if h['analyst_inap'] is not None else 'NA'],
        ['  Confidentiality',                h['analyst_conf'] if h['analyst_conf'] is not None else 'NA'],
        ['Defects (sum of Missed Params)',   h['defects']],
        ['Non-Fatal Defects',                h['nonfatal_defects']],
    ]
    write_table(ws, ['KPI', 'Value'], kpi_rows, start_row=5, freeze=False)

    # MoM
    if metrics.get('mom'):
        m = metrics['mom']
        prev = prev_month_label or 'Prev'
        cur = month_label
        ws.cell(15, 1, "Month over Month").font = section_font
        def _xl_acc(prev_v, cur_v, delta_v):
            p = f"{prev_v:.2f}%" if prev_v is not None else 'NA'
            c = f"{cur_v:.2f}%"  if cur_v  is not None else 'NA'
            d = f"{delta_v:+.2f}%" if delta_v is not None else 'NA'
            return p, c, d
        proc_p, proc_c, proc_d = _xl_acc(m['prev_proc_acc'], m['cur_proc_acc'], m['proc_acc_delta'])
        anl_p,  anl_c,  anl_d  = _xl_acc(m['prev_anl_acc'],  m['cur_anl_acc'],  m['anl_acc_delta'])
        mom_rows = [
            ['Total Audited', m['prev_audited'], m['cur_audited'], f"{m['audited_delta']:+d}"],
            ['Fatal',         f"{m['prev_fatal']} ({m['prev_fatal_pct']:.1f}%)",
                              f"{m['cur_fatal']} ({m['cur_fatal_pct']:.1f}%)", f"{m['fatal_delta']:+d}"],
            ['Process Accuracy', proc_p, proc_c, proc_d],
            ['Analyst Accuracy', anl_p,  anl_c,  anl_d],
        ]
        write_table(ws, ['Metric', prev, cur, 'Delta'], mom_rows, start_row=16, freeze=False)

    # Top Defects
    ws.cell(23, 1, "Top Defect Parameters").font = section_font
    write_table(ws, ['Parameter', 'Defect Count'], metrics['top_defects'], start_row=24, freeze=False)

    # ============ Sheet 2: Performance (combined) ============
    ws2 = wb.create_sheet("Performance")
    ws2.cell(1, 1, "Performance Detail").font = title_font

    # --- Section 2a: Org Scorecard ---
    ws2.cell(3, 1, "Org Scorecard").font = section_font
    org_headers = ['Org','Cases Audited','Accuracy','Fatal','Non-Fatal Defects']
    org_rows = [[o['org'], o['audited'], f"{o['accuracy']:.2f}%", o['fatal'], o['nonfatal']]
                for o in metrics['orgs']]
    write_table(ws2, org_headers, org_rows, start_row=4,
                color_cols={3:('rag',False,False)}, freeze=False)
    org_end = 4 + len(org_rows) + 2

    # --- Section 2b: Analysts ---
    ws2.cell(org_end, 1, "Analyst Performance").font = section_font
    a_headers = ['Analyst','Org','Manager','Audited','Accuracy','Fatal','Non-Fatal Defects']
    a_rows = [[a['analyst'], a['org'], a.get('manager') or '(none)', a['audited'],
               f"{a['accuracy']:.2f}%", a['fatal'], a['nonfatal']] for a in metrics['analysts']]
    write_table(ws2, a_headers, a_rows, start_row=org_end + 1,
                color_cols={5:('rag',False,False)}, freeze=False)
    a_end = org_end + 1 + len(a_rows) + 2

    # --- Section 2c: All Defects (fatal + non-fatal, with comments) ---
    ws2.cell(a_end, 1, "All Defects").font = section_font
    def_rows = []
    for r in metrics['rows']:
        if not r['has_def']:
            continue
        # Determine type
        if r['is_fatal']:
            dtype = 'Fatal'
        elif r['is_nonfatal']:
            dtype = 'Non-Fatal'
        else:
            dtype = 'Mixed'
        def_rows.append([r['case'], r['analyst'], r['org'], r.get('manager') or '(none)',
                         r['date'], dtype, r['missed'],
                         ', '.join(r['params']), r['comment'] or '(no comment)'])
    d_headers = ['Case','Analyst','Org','Manager','Date','Type','# Missed','Parameters','Audit Comment']
    write_table(ws2, d_headers, def_rows, start_row=a_end + 1, freeze=False)

    # ============ Sheet 3: Disputes ============
    ws3 = wb.create_sheet("Disputes")
    ws3.cell(1, 1, "Disputes").font = title_font
    if disputes_df is not None and len(disputes_df) > 0:
        d_rows = []
        for _, r in disputes_df.iterrows():
            d_rows.append([
                r.get('_case', ''),
                r.get('_analyst', ''),
                r.get('_org', ''),
                r.get('_owner', '').title() if r.get('_owner') else '',
                r.get('_report', '').title() if r.get('_report') else '',
                r.get('_category', ''),
                r['_audit_date'].strftime('%Y-%m-%d') if r.get('_audit_date') is not None and pd.notna(r.get('_audit_date')) else '',
                r['_dispute_date'].strftime('%Y-%m-%d') if r.get('_dispute_date') is not None and pd.notna(r.get('_dispute_date')) else '',
                r.get('_workdays', '') if r.get('_workdays') is not None else '',
                r.get('_sla', ''),
            ])
        write_table(ws3, ['Case','Analyst','Org','Defect Owner','Report Info','Category','Audit Date','Dispute Date','Working Days','SLA Status'],
                    d_rows, start_row=3)
    else:
        ws3.cell(3, 1, "No disputes data available.")

    # ============ Sheet 4: Rectification ============
    ws4 = wb.create_sheet("Rectification")
    ws4.cell(1, 1, "Rectification").font = title_font
    if rect_df is not None and len(rect_df) > 0:
        r_rows = []
        for _, r in rect_df.iterrows():
            r_rows.append([
                r.get('_case', ''),
                r.get('_analyst', ''),
                r.get('_org', ''),
                r.get('_status', ''),
                r['_audit_date'].strftime('%Y-%m-%d') if r.get('_audit_date') is not None and pd.notna(r.get('_audit_date')) else '',
                r['_end_date'].strftime('%Y-%m-%d') if r.get('_end_date') is not None and pd.notna(r.get('_end_date')) else '',
                r.get('_workdays', '') if r.get('_workdays') is not None else '',
                r.get('_sla', ''),
            ])
        write_table(ws4, ['Case','Analyst','Org','Status','Audit Date','Rect Date','Working Days','SLA Status'],
                    r_rows, start_row=3)
    else:
        ws4.cell(3, 1, "No rectification data available.")

    # ============ Sheet 5: IVOC ============
    ws5 = wb.create_sheet("IVOC")
    ws5.cell(1, 1, "IVOC").font = title_font
    if ivoc_df is not None and len(ivoc_df) > 0:
        i_rows = []
        for _, r in ivoc_df.iterrows():
            i_rows.append([
                r.get('_case', ''),
                r.get('_analyst', ''),
                r.get('_org', ''),
                r.get('_status', ''),
                r['_audit_date'].strftime('%Y-%m-%d') if r.get('_audit_date') is not None and pd.notna(r.get('_audit_date')) else '',
                r['_end_date'].strftime('%Y-%m-%d') if r.get('_end_date') is not None and pd.notna(r.get('_end_date')) else '',
                r.get('_workdays', '') if r.get('_workdays') is not None else '',
                r.get('_sla', ''),
            ])
        write_table(ws5, ['Case','Analyst','Org','Status','Audit Date','IVOC Date','Working Days','SLA Status'],
                    i_rows, start_row=3)
    else:
        ws5.cell(3, 1, "No IVOC data available.")

    # ============ Sheet 6: Defect Reduction ============
    ws6 = wb.create_sheet("Defect Reduction")
    ws6.cell(1, 1, f"Defect Reduction — {month_label}").font = title_font
    ws6.cell(2, 1, "Filtered to current month only").font = Font(italic=True, color='6B7280')
    if defred_df is not None and len(defred_df) > 0:
        dr_rows = []
        for _, r in defred_df.iterrows():
            dr_rows.append([
                r.get('_case', ''),
                r.get('_analyst', ''),
                r.get('_org', ''),
                r.get('_status', ''),
                r.get('_action', ''),
                r['_review_date'].strftime('%Y-%m-%d') if r.get('_review_date') is not None and pd.notna(r.get('_review_date')) else '',
                r['_update_date'].strftime('%Y-%m-%d') if r.get('_update_date') is not None and pd.notna(r.get('_update_date')) else '',
                r.get('_workdays', '') if r.get('_workdays') is not None else '',
                r.get('_sla', ''),
            ])
        write_table(ws6, ['Case','Analyst','Org','Status','Action Taken','Date of Review','Case Update Date','Working Days','SLA Status'],
                    dr_rows, start_row=4)
    else:
        ws6.cell(4, 1, "No defect reduction data available for this month.")

    wb.save(out_path)
    return out_path


# ============================================================
# SCREENSHOT — matplotlib-based summary image (no external installs)
# ============================================================
def build_screenshot(metrics, process_label, month_label, prev_month_label, png_path):
    """Generate executive briefing PNG: KPIs with MoM, Inflow MoM, Audit Coverage,
    Bottom 3 orgs, Goal banner, Action items. Tuned for email body."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        print("  WARNING: matplotlib not installed; skipping PNG.")
        print("           Install with: python -m pip install matplotlib --user")
        return None

    h = metrics['headline']
    mom = metrics.get('mom')
    coverage = metrics.get('coverage', [])
    inflow = metrics.get('inflow_mom', [])
    goal = metrics.get('defect_reduction_goal') or {}

    # Compute MoM deltas for KPI arrows
    audited_delta = mom['audited_delta'] if mom else 0
    fatal_delta = mom['fatal_delta'] if mom else 0
    proc_acc_delta = mom['proc_acc_delta'] if mom else None

    fig = plt.figure(figsize=(14, 11), facecolor='white')
    fig.suptitle(f"QC Snapshot — {process_label} {month_label}",
                 fontsize=16, fontweight='bold', y=0.97)

    # --- KPI ROW (4 cards with MoM arrows) ---
    def _arrow(delta, lower_is_better=False):
        if delta is None or delta == 0:
            return ''
        if lower_is_better:
            color = '#16A34A' if delta < 0 else '#DC2626'
            arrow = '↓' if delta < 0 else '↑'
        else:
            color = '#16A34A' if delta > 0 else '#DC2626'
            arrow = '↑' if delta > 0 else '↓'
        sign = '+' if delta > 0 else ''
        return f'{arrow} {sign}{delta}', color

    # Build sub-text with MoM trend
    aud_arrow = _arrow(audited_delta, lower_is_better=False) if mom else ''
    fat_arrow = _arrow(fatal_delta, lower_is_better=True) if mom else ''
    acc_arrow = _arrow(proc_acc_delta, lower_is_better=False) if mom and proc_acc_delta is not None else ''

    proc_acc_str = f"{h['accuracy']:.1f}%"
    anl_acc_str = f" / {mom['cur_anl_acc']:.1f}%" if mom and mom.get('cur_anl_acc') is not None else ''

    fatal_str = f"{h['fatal']}"
    if h.get('analyst_fatal') is not None:
        fatal_str += f" / {h['analyst_fatal']}"

    kpi_data = [
        ('CASES AUDITED', str(h['audited']),
            f"vs {mom['prev_audited']} prev" if mom else "From Process file",
            aud_arrow if isinstance(aud_arrow, tuple) else None,
            '#FFFFFF'),
        ('ACCURACY (P / A)', proc_acc_str + anl_acc_str,
            f"vs {mom['prev_proc_acc']:.1f}% prev" if mom else "Average",
            acc_arrow if isinstance(acc_arrow, tuple) else None,
            '#EAF3DE' if h['accuracy'] >= 95 else ('#FAEEDA' if h['accuracy'] >= 90 else '#FCEBEB')),
        ('FATAL (P / A)', fatal_str,
            f"P: I{h['inap']} C{h['conf']}" + (f" • A: I{h['analyst_inap']} C{h['analyst_conf']}" if h.get('analyst_fatal') is not None else ''),
            fat_arrow if isinstance(fat_arrow, tuple) else None,
            '#FCEBEB'),
        ('DEFECTS', str(h['defects']),
            'Total missed params',
            None,
            '#FFFFFF'),
    ]

    for i, (label, value, sub, arrow_tuple, bg) in enumerate(kpi_data):
        ax = fig.add_axes([0.04 + i * 0.235, 0.83, 0.20, 0.11])
        ax.set_facecolor(bg)
        ax.set_xticks([]); ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_edgecolor('#D1D5DB')
        ax.text(0.5, 0.85, label, ha='center', va='center',
                fontsize=9, color='#6B7280', fontweight='bold', transform=ax.transAxes)
        ax.text(0.5, 0.50, value, ha='center', va='center',
                fontsize=20, color='#DC2626' if 'FATAL' in label else '#1F2937',
                fontweight='bold', transform=ax.transAxes)
        ax.text(0.5, 0.18, sub, ha='center', va='center',
                fontsize=7, color='#6B7280', transform=ax.transAxes)
        if arrow_tuple:
            arrow_str, arrow_color = arrow_tuple
            ax.text(0.92, 0.85, arrow_str, ha='right', va='center',
                    fontsize=9, color=arrow_color, fontweight='bold', transform=ax.transAxes)

    # --- INFLOW MoM + AUDIT COVERAGE strip (side by side) ---
    # Inflow MoM (left)
    ax_inflow = fig.add_axes([0.04, 0.66, 0.45, 0.13])
    ax_inflow.set_title("Inflow MoM (from config)", fontsize=10, loc='left', fontweight='bold')
    ax_inflow.axis('off')
    if inflow:
        rows = [['Category', f"{prev_month_label or 'Prev'}", f"{month_label}", 'Δ']]
        for r in inflow:
            d = f"{r['delta_pct']:+.1f}%" if r['delta_pct'] is not None else '—'
            if r['flag']:
                d += ' ⚠'
            rows.append([r['category'], f"{r['prev']:,}", f"{r['cur']:,}", d])
        tbl = ax_inflow.table(cellText=rows[1:], colLabels=rows[0],
                              cellLoc='center', loc='upper left',
                              colWidths=[0.22, 0.22, 0.22, 0.30])
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9)
        tbl.scale(1, 1.4)
        for c in range(4):
            tbl[(0, c)].set_facecolor('#E5E7EB')
            tbl[(0, c)].set_text_props(weight='bold')
        for r_idx, r in enumerate(inflow, 1):
            if r['flag']:
                tbl[(r_idx, 3)].set_facecolor('#FAEEDA')

    # Audit Coverage (right)
    ax_cov = fig.add_axes([0.54, 0.66, 0.42, 0.13])
    ax_cov.set_title("Audit Coverage", fontsize=10, loc='left', fontweight='bold')
    ax_cov.axis('off')
    if coverage:
        rows = [['Category', 'Inflow', 'Audited', '%']]
        for r in coverage:
            cov = f"{r['coverage_pct']:.1f}%" if r['coverage_pct'] is not None else '—'
            rows.append([r['category'], f"{r['total']:,}", f"{r['audited']:,}", cov])
        tbl = ax_cov.table(cellText=rows[1:], colLabels=rows[0],
                           cellLoc='center', loc='upper left',
                           colWidths=[0.22, 0.22, 0.22, 0.22])
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9)
        tbl.scale(1, 1.4)
        for c in range(4):
            tbl[(0, c)].set_facecolor('#E5E7EB')
            tbl[(0, c)].set_text_props(weight='bold')

    # --- DEFECT REDUCTION GOAL BANNER ---
    ax_goal = fig.add_axes([0.04, 0.56, 0.92, 0.07])
    ax_goal.axis('off')
    if goal and goal.get('actual_change_pct') is not None:
        if goal['goal_met']:
            bg_color = '#EAF3DE'
            text_color = '#065F46'
            badge = '✅ MET'
        else:
            bg_color = '#FCEBEB'
            text_color = '#991B1B'
            badge = '❌ NOT MET'
        ax_goal.add_patch(plt.Rectangle((0, 0), 1, 1, facecolor=bg_color,
                                         edgecolor='#D1D5DB', linewidth=1, transform=ax_goal.transAxes))
        ax_goal.text(0.02, 0.5,
                     f"DEFECT REDUCTION GOAL: −{goal['target_pct']:.0f}% target",
                     ha='left', va='center', fontsize=10, fontweight='bold',
                     color=text_color, transform=ax_goal.transAxes)
        ax_goal.text(0.45, 0.5,
                     f"{goal['prev_label']}: {goal['prev_count']} → {goal['cur_label']}: {goal['cur_count']}  •  Actual: {goal['actual_change_pct']:+.1f}%",
                     ha='left', va='center', fontsize=10, color=text_color, transform=ax_goal.transAxes)
        ax_goal.text(0.98, 0.5, badge, ha='right', va='center',
                     fontsize=12, fontweight='bold', color=text_color, transform=ax_goal.transAxes)

    # --- TOP DEFECTS (left) + BOTTOM 3 ORGS (right) ---
    ax_td = fig.add_axes([0.04, 0.23, 0.45, 0.30])
    ax_td.set_title("Top Defect Parameters", fontsize=10, loc='left', fontweight='bold')
    if metrics['top_defects']:
        names = [n[:30] for n, _ in metrics['top_defects'][:7]]
        counts = [c for _, c in metrics['top_defects'][:7]]
        ax_td.barh(range(len(names)), counts, color='#BA7517', edgecolor='white')
        ax_td.set_yticks(range(len(names)))
        ax_td.set_yticklabels(names, fontsize=8)
        ax_td.invert_yaxis()
        for i, c in enumerate(counts):
            ax_td.text(c, i, f' {c}', va='center', fontsize=8)
    ax_td.tick_params(axis='x', labelsize=7)
    for spine in ['top', 'right']:
        ax_td.spines[spine].set_visible(False)

    # Bottom Performers - bottom 8 analysts by accuracy (Org + Analyst Name only)
    ax_org = fig.add_axes([0.54, 0.23, 0.42, 0.30])
    ax_org.set_title("Bottom Performers", fontsize=10, loc='left', fontweight='bold')
    ax_org.axis('off')
    if metrics.get('analysts'):
        bot_analysts = sorted(metrics['analysts'], key=lambda x: x['accuracy'])[:8]
        rows = [['Org', 'Analyst']]
        for a in bot_analysts:
            rows.append([a['org'], a['analyst']])
        tbl = ax_org.table(cellText=rows[1:], colLabels=rows[0],
                           cellLoc='center', loc='upper left',
                           colWidths=[0.25, 0.55])
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9)
        tbl.scale(1, 1.4)
        for c in range(2):
            tbl[(0, c)].set_facecolor('#E5E7EB')
            tbl[(0, c)].set_text_props(weight='bold')

    # --- ACTION ITEMS strip ---
    actions = []
    # 1. Inflow spikes
    for r in inflow or []:
        if r['flag']:
            direction = 'up' if r['delta_pct'] and r['delta_pct'] > 0 else 'down'
            actions.append(f"• {r['category']} volume {direction} {abs(r['delta_pct']):.0f}% MoM — investigate root cause")
    # 2. Goal status
    if goal and goal.get('actual_change_pct') is not None:
        if goal['goal_met']:
            actions.append(f"• Defect reduction goal MET ({goal['actual_change_pct']:+.1f}% vs −{goal['target_pct']:.0f}% target)")
        else:
            actions.append(f"• Defect reduction goal MISSED ({goal['actual_change_pct']:+.1f}% vs −{goal['target_pct']:.0f}% target) — focus coaching")
    # 3. Bottom org
    if metrics['orgs']:
        worst = sorted(metrics['orgs'], key=lambda x: x['accuracy'])[0]
        if worst['accuracy'] < 90:
            actions.append(f"• {worst['org']} accuracy {worst['accuracy']:.1f}% — below 90% threshold")
    # 4. Pending — Missed SLA across status cards
    breached_total = 0
    for key in ('rectification', 'ivoc', 'defect_reduction'):
        sc = metrics.get(key) or {}
        breakdown = sc.get('breakdown') or {}
        breached_total += breakdown.get('Pending — Missed SLA', 0)
    if breached_total > 0:
        actions.append(f"• {breached_total} cases missed SLA across Rectification/IVOC/Defect Reduction — ops follow-up")

    # 5. Analyst Pareto - top contributors to defects
    if metrics.get('analysts'):
        # Compute defect contribution per analyst (fatal + non-fatal)
        analysts_with_def = [{'name': a['analyst'], 'org': a['org'],
                              'def_count': a['fatal'] + a['nonfatal']}
                             for a in metrics['analysts']]
        analysts_with_def = [a for a in analysts_with_def if a['def_count'] > 0]
        total_def_count = sum(a['def_count'] for a in analysts_with_def)
        if total_def_count > 0 and len(analysts_with_def) >= 5:
            top_n = 8 if len(analysts_with_def) >= 10 else 5
            top_analysts = sorted(analysts_with_def, key=lambda x: -x['def_count'])[:top_n]
            top_count = sum(a['def_count'] for a in top_analysts)
            top_pct = top_count / total_def_count * 100
            actions.append(f"• Top {top_n} analysts account for {top_pct:.0f}% of defects ({top_count} of {total_def_count}) — coaching priority")

    ax_action = fig.add_axes([0.04, 0.04, 0.92, 0.16])
    ax_action.axis('off')
    ax_action.text(0, 0.95, "ACTION ITEMS", ha='left', va='top', fontsize=10, fontweight='bold',
                   color='#1F2937', transform=ax_action.transAxes)
    if actions:
        for i, a in enumerate(actions[:5]):
            ax_action.text(0, 0.78 - i * 0.18, a, ha='left', va='top', fontsize=9,
                           color='#374151', transform=ax_action.transAxes)
    else:
        ax_action.text(0, 0.78, "No critical action items flagged this period.",
                       ha='left', va='top', fontsize=9, color='#6B7280', transform=ax_action.transAxes)

    plt.savefig(png_path, dpi=120, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    return png_path


# ============================================================
# EMAIL BODY BUILDER (Outlook-paste-friendly)
# ============================================================
def build_email_body(metrics, process_label, month_label, prev_month_label, dashboard_url, html_path):
    """Generate email-body HTML file. Inline CSS, table-based, no JS.
    User opens in browser, Ctrl+A, Ctrl+C, paste into Outlook."""
    h = metrics['headline']
    mom = metrics.get('mom') or {}
    goal = metrics.get('defect_reduction_goal') or {}
    top_defects = metrics.get('top_defects', [])
    impact = metrics.get('impact_categorization', [])
    orgs = metrics.get('orgs', [])

    # Build pieces
    # KPI table
    kpi_rows = ''
    kpi_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;font-weight:500;">Cases Audited</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{h["audited"]}</td></tr>'
    acc_str = f"{h['accuracy']:.2f}%"
    if mom.get('cur_anl_acc') is not None:
        acc_str += f" (Process) / {mom['cur_anl_acc']:.2f}% (Analyst)"
    kpi_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;font-weight:500;">Accuracy</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{acc_str}</td></tr>'
    fatal_str = f"{h['fatal']} (Inap {h['inap']}, Conf {h['conf']})"
    if h.get('analyst_fatal') is not None:
        fatal_str += f" / Analyst: {h['analyst_fatal']}"
    kpi_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;font-weight:500;">Fatal Errors</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{fatal_str}</td></tr>'
    kpi_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;font-weight:500;">Total Defects</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{h["defects"]}</td></tr>'
    if goal.get('actual_change_pct') is not None:
        badge = '✅ MET' if goal['goal_met'] else '❌ NOT MET'
        kpi_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;background:#F1EFE8;font-weight:500;">Defect Reduction Goal</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{goal["actual_change_pct"]:+.1f}% vs −{goal["target_pct"]:.0f}% target — {badge}</td></tr>'

    # MoM table
    mom_table = ''
    if mom:
        mom_table = f'''
<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;margin:0 0 20px 0;">
  <thead><tr style="background:#F1EFE8;">
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">Metric</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('prev_label','Prev')}</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('cur_label','Cur')}</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">Δ</th>
  </tr></thead>
  <tbody>
    <tr><td style="padding:8px 12px;border:1px solid #D3D1C7;">Cases Audited</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('prev_audited',0)}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('cur_audited',0)}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('audited_delta',0):+d}</td></tr>
    <tr><td style="padding:8px 12px;border:1px solid #D3D1C7;">Process Accuracy</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('prev_proc_acc',0):.2f}%</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('cur_proc_acc',0):.2f}%</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('proc_acc_delta',0):+.2f}</td></tr>
    <tr><td style="padding:8px 12px;border:1px solid #D3D1C7;">Analyst Accuracy</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('prev_anl_acc',0):.2f}%</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('cur_anl_acc',0):.2f}%</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('anl_acc_delta',0):+.2f}</td></tr>
    <tr><td style="padding:8px 12px;border:1px solid #D3D1C7;">Fatal Errors</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('prev_fatal',0)}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('cur_fatal',0)}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{mom.get('fatal_delta',0):+d}</td></tr>
  </tbody>
</table>'''

    # Top 3 defects (table with placeholder for causes)
    top3_rows = ''
    for name, cnt in top_defects[:3]:
        pct = (cnt / h['defects'] * 100) if h['defects'] else 0
        top3_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;">{name}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{cnt}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{pct:.1f}%</td><td style="padding:8px 12px;border:1px solid #D3D1C7;color:#5F5E5A;font-style:italic;">[Fill potential causes]</td></tr>'

    # Impact categorization
    impact_rows = ''
    for r in impact:
        impact_rows += f'<tr><td style="padding:8px 12px;border:1px solid #D3D1C7;">{r["category"]}</td><td style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">{r["defects"]}</td></tr>'

    # Final HTML
    html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>QC Email Body — {process_label} {month_label}</title></head>
<body style="font-family:Arial,sans-serif;color:#2C2C2A;max-width:760px;margin:20px auto;padding:20px;background:#FAF8F4;font-size:14px;line-height:1.5;">

<p style="color:#5F5E5A;font-size:12px;font-style:italic;border:1px dashed #D3D1C7;padding:8px;background:#FFFFFF;">📋 <strong>Instructions:</strong> Open in browser → press Ctrl+A → Ctrl+C → paste into Outlook (HTML format). Add your Executive Summary and Page 0 link manually before sending.</p>

<p style="color:#791F1F;font-style:italic;">&lt;Please do not forward&gt;</p>

<p>Dear Team,</p>

<p>Please find the <strong>{month_label} QC Audit Metrics for {process_label}</strong> below. The complete metrics, drill-downs, and case-level data are available in the interactive dashboard linked at the end of this email.</p>

<p>📄 Refer to <em>Page 0</em> for definitions and methodology: <strong>[your link]</strong></p>

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Executive Summary</h3>
<ul>
  <li>[April performance mixed. Process accuracy at X%, Analyst accuracy at X%]</li>
  <li>[Reopen volume +X% MoM — material inflow shift; recommend ops root-cause review]</li>
  <li>[Top defects: Defect 1 and Defect 2 contributed X% of total defects — concentrated in orgs]</li>
  <li>[Defect Reduction goal MET / NOT MET — X% reduction vs −10% target]</li>
  <li>[Org 1, Org 2 below 95% threshold — coaching pipeline active for May]</li>
</ul>

<p style="color:#5F5E5A;font-size:12px;font-style:italic;">Detailed KPIs, MoM comparisons, org/analyst breakdowns, and SLA status are available in the dashboard (link below).</p>

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Key Metrics — {month_label}</h3>
<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;margin:0 0 20px 0;">
  <tbody>{kpi_rows}</tbody>
</table>

<h3 style="color:#2C2C2A;font-weight:500;">Month-over-Month Comparison</h3>
{mom_table}

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Top 3 Defects — Drivers and Potential Causes</h3>
<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;margin:0 0 20px 0;">
  <thead><tr style="background:#F1EFE8;">
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">Parameter</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">Defects</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">Contribution %</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">Potential Causes</th>
  </tr></thead>
  <tbody>{top3_rows}</tbody>
</table>

<p style="color:#5F5E5A;font-size:12px;font-style:italic;">For full defect parameter breakdown and case-level drill-downs, refer to the dashboard.</p>

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Defect Impact Categorization</h3>
<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;margin:0 0 20px 0;">
  <thead><tr style="background:#F1EFE8;">
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">Category</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:right;">Defects</th>
  </tr></thead>
  <tbody>{impact_rows}</tbody>
</table>

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Trend Analysis — Dec'25 to {month_label}</h3>
<p style="color:#5F5E5A;font-style:italic;">[Insert multi-month trending table from past audit cycles. Manual entry.]</p>

<p style="color:#5F5E5A;font-size:12px;font-style:italic;">For real-time MoM and WoW views with org-level filtering, refer to the dashboard.</p>

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Interactive QC Dashboard</h3>
<p><strong>🔗 Access:</strong> <a href="{dashboard_url}" style="color:#BA7517;">{dashboard_url}</a></p>

<p>The dashboard provides the complete view of this month's metrics and is structured for self-service use. Key capabilities:</p>
<ul>
  <li>Live filtering across org, analyst, week, audit category, and defect topic</li>
  <li>Drill-down on any KPI, chart, or table row to view underlying cases with audit commentary</li>
  <li>Section-level CSV export for ad-hoc analysis</li>
  <li>SLA tracking for Disputes, Rectification, IVOC, and Defect Reduction</li>
  <li>Defect Impact Categorization mapping each defect to a business-impact category</li>
  <li>Defect Reduction Goal tracker against the 10% month-over-month target</li>
</ul>

<p>For any specific cuts of the data or further drill-downs, please respond to this email.</p>

<hr style="border:none;border-top:1px solid #D3D1C7;margin:20px 0;"/>

<h3 style="color:#2C2C2A;font-weight:500;">Calibration Cycle</h3>
<table style="border-collapse:collapse;width:100%;font-family:Arial,sans-serif;font-size:13px;margin:0 0 20px 0;">
  <thead><tr style="background:#F1EFE8;">
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">Calibration with Ops Team</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">Audit Parameter Revision</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">QC SPOC</th>
    <th style="padding:8px 12px;border:1px solid #D3D1C7;text-align:left;">QC Manager</th>
  </tr></thead>
  <tbody><tr>
    <td style="padding:8px 12px;border:1px solid #D3D1C7;color:#5F5E5A;">[Last / Next]</td>
    <td style="padding:8px 12px;border:1px solid #D3D1C7;color:#5F5E5A;">[Last / Next]</td>
    <td style="padding:8px 12px;border:1px solid #D3D1C7;color:#5F5E5A;">[SPOC names]</td>
    <td style="padding:8px 12px;border:1px solid #D3D1C7;color:#5F5E5A;">[Manager LDAP]</td>
  </tr></tbody>
</table>

<p>Regards,<br/>
Niki<br/>
QC | Retail FinCoM<br/>
Amazon</p>

</body></html>'''

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return html_path


# ============================================================
# MAIN ORCHESTRATION
# ============================================================
def list_datasets():
    if not BASE_FOLDER.exists():
        print(f"\nERROR: Folder not found: {BASE_FOLDER}")
        sys.exit(1)
    return sorted([f for f in BASE_FOLDER.iterdir() if f.is_dir()], key=lambda p: p.name)


def parse_folder_name(name):
    """Parse 'Process_Month_Year' (e.g. 'General_Mar_2026') into (process, month, year)."""
    parts = name.split('_')
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    return name, '', ''


def find_prev_month_folder(current_folder):
    """Given e.g. 'General_Mar_2026', find 'General_Feb_2026' if it exists."""
    process, month, year = parse_folder_name(current_folder.name)
    if month not in MONTHS_ORDER:
        return None, None
    idx = MONTHS_ORDER.index(month)
    if idx == 0:
        prev_month = "Dec"
        prev_year = str(int(year) - 1) if year.isdigit() else year
    else:
        prev_month = MONTHS_ORDER[idx - 1]
        prev_year = year
    candidate = BASE_FOLDER / f"{process}_{prev_month}_{prev_year}"
    if candidate.exists() and candidate.is_dir():
        return candidate, f"{prev_month} {prev_year}"
    return None, None


def run_dataset(folder):
    process, month, year = parse_folder_name(folder.name)
    process_label = process
    month_label = f"{month} {year}".strip()

    print(f"\n{'=' * 60}")
    print(f"Running: {process_label} — {month_label}")
    print('=' * 60)

    print("\nLoading config...")
    config = load_config(folder)

    # Parse current month's process & analyst files
    print("Reading current month files...")
    process_df = parse_audit_file(folder / "Fincom_Process.csv", "Process Audit",
                                   parameters=config.get('parameters'),
                                   inap_pos=int(config.get('fatal_inap_position', 4)),
                                   conf_pos=int(config.get('fatal_conf_position', 9)))
    analyst_df = parse_audit_file(folder / "Fincom_Analyst.csv", "Analyst Audit",
                                   parameters=config.get('parameters'),
                                   inap_pos=int(config.get('fatal_inap_position', 4)),
                                   conf_pos=int(config.get('fatal_conf_position', 9)))

    if len(process_df) == 0:
        print("\nERROR: No data in Process file. Cannot continue.")
        return

    # Other files
    # Track which files exist (for "missing -> skip card" logic)
    files_present = {
        'disputes':         (folder / "Disputes.csv").exists(),
        'rectification':    (folder / "Rectification.csv").exists(),
        'ivoc':             (folder / "IVOC.csv").exists(),
        'defect_reduction': (folder / "Defect Reduction.csv").exists(),
    }

    disputes_df = parse_disputes(folder / "Disputes.csv", int(config['disputes_sla_days']))
    rect_df     = parse_rectification(folder / "Rectification.csv", int(config['rectification_sla_days']))
    ivoc_df     = parse_ivoc(folder / "IVOC.csv", int(config['ivoc_sla_days']))
    defred_df   = parse_defect_reduction(folder / "Defect Reduction.csv", month, int(config['reduction_sla_days']))

    # Track missing
    missing = []
    if not (folder / "Rectification.csv").exists(): missing.append("Rectification")
    if not (folder / "IVOC.csv").exists():         missing.append("IVOC")
    if not (folder / "Defect Reduction.csv").exists(): missing.append("Defect Reduction")

    # Previous month for MoM
    prev_folder, prev_month_label = find_prev_month_folder(folder)
    prev_process_df = pd.DataFrame()
    prev_analyst_df = pd.DataFrame()
    prev_config = None
    if prev_folder:
        print(f"\nPrevious month folder found: {prev_folder.name}")
        prev_config_loaded = load_config(prev_folder)
        prev_process_df = parse_audit_file(prev_folder / "Fincom_Process.csv", "Process Audit (prev)",
                                            parameters=prev_config_loaded.get('parameters'),
                                            inap_pos=int(prev_config_loaded.get('fatal_inap_position', 4)),
                                            conf_pos=int(prev_config_loaded.get('fatal_conf_position', 9)))
        prev_analyst_df = parse_audit_file(prev_folder / "Fincom_Analyst.csv", "Analyst Audit (prev)",
                                            parameters=prev_config_loaded.get('parameters'),
                                            inap_pos=int(prev_config_loaded.get('fatal_inap_position', 4)),
                                            conf_pos=int(prev_config_loaded.get('fatal_conf_position', 9)))
        prev_config = prev_config_loaded
    else:
        print("\nNo previous month folder found — MoM section will be hidden.")

    # Compute metrics
    print("\nComputing metrics...")
    metrics = compute_metrics(process_df, analyst_df, disputes_df, rect_df, ivoc_df, defred_df,
                              prev_process_df, prev_analyst_df, config, prev_config,
                              process_label, month_label, prev_month_label)
    if metrics is None:
        print("ERROR: Could not compute metrics.")
        return

    # Track which files were present (for skip-if-missing logic)
    metrics['files_present'] = files_present

    # Print sanity check
    h = metrics['headline']
    print(f"\n--- HEADLINE NUMBERS (for verification) ---")
    print(f"  Cases Audited:        {h['audited']}")
    print(f"  Accuracy:             {h['accuracy']:.2f}%")
    print(f"  Fatal (Process): {h['fatal']}  (= Inap {h['inap']} + Conf {h['conf']})")
    if h.get('analyst_fatal') is not None:
        print(f"  Fatal (Analyst): {h['analyst_fatal']}  (= Inap {h['analyst_inap']} + Conf {h['analyst_conf']})")
    print(f"  Defects (Σ Missed):   {h['defects']}")
    print(f"  Non-Fatal Defects:    {h['nonfatal_defects']}")

    # Build outputs
    print("\nBuilding HTML dashboard...")
    html = build_html(metrics, process_label, month_label, prev_month_label, missing)
    html_path = folder / f"QC_Dashboard_{folder.name}.html"
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  Saved: {html_path}")

    print("Building Excel report...")
    xlsx_path = folder / f"QC_Report_{folder.name}.xlsx"
    if build_excel(metrics, process_label, month_label, prev_month_label, xlsx_path,
                   disputes_df, rect_df, ivoc_df, defred_df):
        print(f"  Saved: {xlsx_path}")

    print("Building snapshot PNG...")
    png_path = folder / f"QC_Snapshot_{folder.name}.png"
    if build_screenshot(metrics, process_label, month_label, prev_month_label, png_path):
        print(f"  Saved: {png_path}")

    print("\n" + "=" * 60)
    print("ALL DONE")
    print("=" * 60)


def main():
    print("FinCom QC Automation — v4")
    folders = list_datasets()
    if not folders:
        print("No process folders found in", BASE_FOLDER)
        sys.exit(1)
    print("\nAvailable datasets:")
    for i, f in enumerate(folders, 1):
        print(f"  {i}. {f.name}")
    try:
        choice = int(input("\nEnter number to select: ").strip())
        if choice < 1 or choice > len(folders):
            print("Invalid choice.")
            sys.exit(1)
    except (ValueError, KeyboardInterrupt):
        print("\nCancelled.")
        sys.exit(0)
    run_dataset(folders[choice - 1])


if __name__ == "__main__":
    main()
